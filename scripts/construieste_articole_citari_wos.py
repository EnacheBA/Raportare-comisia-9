"""Separă articolele și citările WoS de restul înregistrărilor Google Scholar.

Modulul este conceput pentru a fi importat dintr-un Jupyter Notebook, dar poate fi
rulat și direct din linia de comandă. Fișierele sursă nu sunt modificate.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WOS_HEADERS = [
    "Tip",
    "Autori",
    "Titlu",
    "Revista / Conferința",
    "Data și locul conferinței",
    "ISSN/eISSN",
    "Anul",
    "Volum",
    "Număr",
    "Pagini / nr. articol",
    "DOI",
    "Valoare WOS",
    "Link",
]


def _log(message: str, verbose: bool = True) -> None:
    if verbose:
        print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {message}", flush=True)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def _ascii(value: Any) -> str:
    return (
        unicodedata.normalize("NFKD", _clean(value))
        .encode("ascii", "ignore")
        .decode("ascii")
    )


def _normal_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _ascii(value).lower()).strip()


def _normal_title(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _ascii(value).lower()).strip()


def _normal_doi(value: Any) -> str:
    text = _clean(value).lower()
    text = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", text)
    text = re.sub(r"^doi\s*:\s*", "", text)
    match = re.search(r"10\.\d{4,9}/\S+", text)
    if match:
        text = match.group(0)
    return text.rstrip(".,;:)]}")


def _valid_url(value: Any) -> str:
    text = _clean(value)
    if text in {"", "0", "0.0"}:
        return ""
    if re.match(r"^https?://", text, flags=re.IGNORECASE):
        return text
    return ""


def _record_value(row: dict[str, Any], *aliases: str) -> str:
    normalized = {_normal_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = _clean(normalized.get(_normal_header(alias)))
        if value:
            return value
    return ""


def _join_unique(parts: Iterable[Any], separator: str = "; ") -> str:
    result: list[str] = []
    seen: set[str] = set()
    for part in parts:
        text = _clean(part)
        marker = text.casefold()
        if text and marker not in seen:
            result.append(text)
            seen.add(marker)
    return separator.join(result)


def _normalize_type(value: Any) -> str:
    text = _clean(value).upper()
    if text in {"J", "JOURNAL", "ARTICLE"}:
        return "J"
    if text in {"C", "CONFERENCE", "PROCEEDINGS", "PROCEEDINGS PAPER"}:
        return "C"
    return ""


def _infer_type(row: dict[str, Any]) -> str:
    direct = _normalize_type(
        _record_value(row, "Tip", "Type", "Publication Type")
    )
    if direct:
        return direct
    document_type = _record_value(row, "Document Type")
    conference = _join_unique(
        [
            _record_value(row, "Conference Title"),
            _record_value(row, "Conference Date"),
            _record_value(row, "Conference Location"),
        ]
    )
    combined = f"{document_type} {conference}".lower()
    if conference or re.search(r"conference|symposium|proceedings", combined):
        return "C"
    if re.search(r"\b(article|review|journal)\b", document_type.lower()):
        return "J"
    return ""


def _page_or_article_number(row: dict[str, Any]) -> str:
    existing = _record_value(
        row,
        "Pagini / nr. articol",
        "Pagini",
        "Pages",
        "Page range",
    )
    start = _record_value(row, "Start Page")
    end = _record_value(row, "End Page")
    article_number = _record_value(row, "Article Number")
    number_of_pages = _record_value(row, "Number of Pages")

    if existing:
        page_text = existing
    elif start and end and start != end:
        page_text = f"{start}–{end}"
    else:
        page_text = start or end

    parts: list[str] = []
    if page_text:
        parts.append(page_text)
    if article_number and article_number.casefold() not in page_text.casefold():
        parts.append(f"nr. articol {article_number}")
    if not parts and number_of_pages:
        parts.append(f"{number_of_pages} pagini")
    return _join_unique(parts)


def _canonical_from_wos(row: dict[str, Any]) -> dict[str, str]:
    tip = _normalize_type(
        _record_value(row, "Tip", "Type", "Publication Type")
    )
    if not tip:
        tip = _infer_type(row)

    source_title = _record_value(
        row, "Revista / Conferința", "Source Title", "Journal", "Venue"
    )
    conference_title = _record_value(row, "Conference Title")
    venue = conference_title if tip == "C" and conference_title else source_title

    issn = _record_value(row, "ISSN")
    eissn = _record_value(row, "eISSN")
    combined_issn = _record_value(row, "ISSN/eISSN")
    if not combined_issn:
        labels: list[str] = []
        if issn:
            labels.append(f"ISSN {issn}")
        if eissn and eissn.casefold() != issn.casefold():
            labels.append(f"eISSN {eissn}")
        combined_issn = _join_unique(labels)

    doi = _record_value(row, "DOI")
    link = _valid_url(
        _record_value(
            row,
            "Link",
            "Link articol",
            "DOI Link",
            "Web of Science Record",
        )
    )
    if not link and _normal_doi(doi):
        link = f"https://doi.org/{_normal_doi(doi)}"

    conference_details = _record_value(row, "Data și locul conferinței")
    if not conference_details:
        conference_details = _join_unique(
            [
                _record_value(row, "Conference Date"),
                _record_value(row, "Conference Location"),
            ]
        )

    return {
        "Tip": tip,
        "Autori": _record_value(row, "Autori", "Authors", "Author Full Names"),
        "Titlu": _record_value(row, "Titlu", "Article Title", "Title"),
        "Revista / Conferința": venue,
        "Data și locul conferinței": conference_details,
        "ISSN/eISSN": combined_issn,
        "Anul": _record_value(row, "Anul", "Publication Year", "Year"),
        "Volum": _record_value(row, "Volum", "Volume"),
        "Număr": _record_value(row, "Număr", "Issue", "Number"),
        "Pagini / nr. articol": _page_or_article_number(row),
        "DOI": doi,
        "Valoare WOS": _record_value(
            row, "Valoare WOS", "UT (Unique WOS ID)", "Web of Science Index"
        ),
        "Link": link,
    }


def _canonical_from_google(row: dict[str, Any]) -> dict[str, str]:
    return {
        "Tip": _normalize_type(_record_value(row, "Tip", "Type")),
        "Autori": _record_value(row, "Autori", "Authors"),
        "Titlu": _record_value(row, "Titlu", "Title", "Article Title"),
        "Revista / Conferința": _record_value(
            row, "Revista / Conferința", "Journal / Conference", "Source Title"
        ),
        "Data și locul conferinței": _record_value(
            row, "Data și locul conferinței"
        ),
        "ISSN/eISSN": _record_value(row, "ISSN/eISSN", "ISSN", "eISSN"),
        "Anul": _record_value(row, "Anul", "Year", "Publication Year"),
        "Volum": _record_value(row, "Volum", "Volume"),
        "Număr": _record_value(row, "Număr", "Issue", "Number"),
        "Pagini / nr. articol": _record_value(
            row, "Pagini / nr. articol", "Pagini", "Pages", "Article Number"
        ),
        "DOI": _record_value(row, "DOI"),
        "Valoare WOS": _record_value(row, "Valoare WOS"),
        "Link": _record_value(row, "Link", "Link articol", "Article Link"),
    }


def _merge_records(primary: dict[str, str], fallback: dict[str, str]) -> dict[str, str]:
    merged = {header: _clean(primary.get(header)) or _clean(fallback.get(header)) for header in WOS_HEADERS}
    if merged["Tip"] not in {"J", "C"}:
        merged["Tip"] = fallback.get("Tip", "") if fallback.get("Tip") in {"J", "C"} else ""
    return merged


def _year_compatible(left: dict[str, str], right: dict[str, str]) -> bool:
    year_left = re.search(r"\b(?:19|20)\d{2}\b", _clean(left.get("Anul")))
    year_right = re.search(r"\b(?:19|20)\d{2}\b", _clean(right.get("Anul")))
    return not year_left or not year_right or year_left.group(0) == year_right.group(0)


def _title_similarity(left: Any, right: Any) -> float:
    a, b = _normal_title(left), _normal_title(right)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    ratio = SequenceMatcher(None, a, b).ratio()
    shorter, longer = sorted((a, b), key=len)
    if len(shorter) >= 55 and longer.startswith(shorter):
        ratio = max(ratio, 0.94)
    return ratio


@dataclass
class MatchItem:
    record: dict[str, str]
    payload: Any


class RecordMatcher:
    def __init__(self, items: Iterable[MatchItem]):
        self.items = list(items)
        self.by_doi: dict[str, list[MatchItem]] = {}
        self.by_title: dict[str, list[MatchItem]] = {}
        for item in self.items:
            doi = _normal_doi(item.record.get("DOI"))
            title = _normal_title(item.record.get("Titlu"))
            if doi:
                self.by_doi.setdefault(doi, []).append(item)
            if title:
                self.by_title.setdefault(title, []).append(item)

    def find(self, query: dict[str, str], fuzzy_threshold: float = 0.94) -> MatchItem | None:
        doi = _normal_doi(query.get("DOI"))
        if doi and doi in self.by_doi:
            return self.by_doi[doi][0]

        title = _normal_title(query.get("Titlu"))
        if title and title in self.by_title:
            candidates = self.by_title[title]
            for candidate in candidates:
                if _year_compatible(query, candidate.record):
                    return candidate
            return candidates[0]

        if not title:
            return None
        best: MatchItem | None = None
        best_score = fuzzy_threshold
        for candidate in self.items:
            if not _year_compatible(query, candidate.record):
                continue
            score = _title_similarity(query.get("Titlu"), candidate.record.get("Titlu"))
            if score >= best_score:
                best = candidate
                best_score = score
        return best


class DuplicateTracker:
    def __init__(self) -> None:
        self.dois: set[str] = set()
        self.titles: set[str] = set()
        self.fallbacks: set[str] = set()

    @staticmethod
    def _fallback(record: dict[str, str]) -> str:
        return "|".join(
            _normal_title(record.get(field))
            for field in ("Autori", "Titlu", "Revista / Conferința", "Anul", "Link")
        )

    def contains(self, record: dict[str, str]) -> bool:
        doi = _normal_doi(record.get("DOI"))
        title = _normal_title(record.get("Titlu"))
        fallback = self._fallback(record)
        return bool(
            (doi and doi in self.dois)
            or (title and title in self.titles)
            or (not doi and not title and fallback and fallback in self.fallbacks)
        )

    def add(self, record: dict[str, str]) -> None:
        doi = _normal_doi(record.get("DOI"))
        title = _normal_title(record.get("Titlu"))
        fallback = self._fallback(record)
        if doi:
            self.dois.add(doi)
        if title:
            self.titles.add(title)
        if fallback:
            self.fallbacks.add(fallback)


@dataclass
class GoogleRow:
    row_number: int
    values: dict[str, Any]
    is_article: bool
    canonical: dict[str, str]


@dataclass
class GoogleGroup:
    parent: GoogleRow | None
    citations: list[GoogleRow]


@dataclass
class GoogleData:
    sheet_name: str
    headers: list[str]
    rows: list[GoogleRow]
    groups: list[GoogleGroup]


def _read_google_xlsx(path: Path) -> GoogleData:
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheet = workbook.active
    headers = [_clean(cell.value) for cell in sheet[1]]
    if not any(headers):
        raise ValueError(f"Fișierul Google nu are antete: {path}")

    title_index = next(
        (
            index
            for index, header in enumerate(headers, start=1)
            if _normal_header(header) in {_normal_header("Titlu"), _normal_header("Article Title"), _normal_header("Title")}
        ),
        None,
    )
    if title_index is None:
        raise ValueError("Nu am găsit coloana de titlu în fișierul Articole_Google.")

    link_headers = {
        _normal_header("Link"),
        _normal_header("Link articol"),
        _normal_header("Article Link"),
    }
    rows: list[GoogleRow] = []
    for row_number in range(2, sheet.max_row + 1):
        values: dict[str, Any] = {}
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row_number, column)
            value = cell.value
            if _normal_header(header) in link_headers and cell.hyperlink and cell.hyperlink.target:
                value = cell.hyperlink.target
            values[header] = value
        if not any(_clean(value) for value in values.values()):
            continue
        is_article = bool(sheet.cell(row_number, title_index).font.bold)
        rows.append(
            GoogleRow(
                row_number=row_number,
                values=values,
                is_article=is_article,
                canonical=_canonical_from_google(values),
            )
        )

    groups = _group_google_rows(rows)
    return GoogleData(sheet.title, headers, rows, groups)


def _read_google_xls(path: Path) -> GoogleData:
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError(
            "Pentru fișiere .xls este necesar pachetul xlrd. Rulați: %pip install xlrd"
        ) from exc

    book = xlrd.open_workbook(str(path), formatting_info=True)
    sheet = book.sheet_by_index(0)
    headers = [_clean(sheet.cell_value(0, column)) for column in range(sheet.ncols)]
    title_index = next(
        (
            index
            for index, header in enumerate(headers)
            if _normal_header(header) in {_normal_header("Titlu"), _normal_header("Article Title"), _normal_header("Title")}
        ),
        None,
    )
    if title_index is None:
        raise ValueError("Nu am găsit coloana de titlu în fișierul Articole_Google.")

    rows: list[GoogleRow] = []
    for row_index in range(1, sheet.nrows):
        values = {
            headers[column]: sheet.cell_value(row_index, column)
            for column in range(sheet.ncols)
        }
        if not any(_clean(value) for value in values.values()):
            continue
        xf = book.xf_list[sheet.cell_xf_index(row_index, title_index)]
        is_article = bool(book.font_list[xf.font_index].bold)
        rows.append(
            GoogleRow(
                row_number=row_index + 1,
                values=values,
                is_article=is_article,
                canonical=_canonical_from_google(values),
            )
        )
    groups = _group_google_rows(rows)
    return GoogleData(sheet.name, headers, rows, groups)


def _group_google_rows(rows: list[GoogleRow]) -> list[GoogleGroup]:
    groups: list[GoogleGroup] = []
    current: GoogleGroup | None = None
    for row in rows:
        if row.is_article:
            current = GoogleGroup(parent=row, citations=[])
            groups.append(current)
        elif current is not None:
            current.citations.append(row)
        else:
            current = GoogleGroup(parent=None, citations=[row])
            groups.append(current)
    return groups


def read_google(path: str | Path) -> GoogleData:
    source = Path(path).expanduser().resolve()
    if source.suffix.lower() == ".xls":
        return _read_google_xls(source)
    return _read_google_xlsx(source)


def read_table(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path).expanduser().resolve()
    try:
        if source.suffix.lower() == ".xls":
            dataframe = pd.read_excel(source, dtype=object, engine="xlrd")
        else:
            dataframe = pd.read_excel(source, dtype=object, engine="openpyxl")
    except ImportError as exc:
        raise ImportError(
            "Lipsește o bibliotecă pentru citirea Excel. Rulați în notebook: "
            "%pip install pandas openpyxl xlrd"
        ) from exc
    dataframe = dataframe.where(pd.notna(dataframe), "")
    return dataframe.to_dict(orient="records")


def _column_widths(sheet, headers: list[str]) -> None:
    for column, header in enumerate(headers, start=1):
        longest = len(header)
        for row in range(2, min(sheet.max_row, 300) + 1):
            longest = max(longest, len(_clean(sheet.cell(row, column).value)))
        if header in {"Autori", "Titlu", "Revista / Conferința"}:
            width = min(max(longest + 2, 24), 52)
        elif "Link" in header:
            width = 38
        else:
            width = min(max(longest + 2, 11), 28)
        sheet.column_dimensions[get_column_letter(column)].width = width


def _apply_common_sheet_format(sheet, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    _column_widths(sheet, headers)


def _write_wos(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Articole și citări WoS"
    sheet.append(WOS_HEADERS)
    article_fill = PatternFill("solid", fgColor="D9EAF7")

    for entry in rows:
        record = entry["record"]
        is_article = bool(entry["is_article"])
        values: list[Any] = []
        for header in WOS_HEADERS:
            value: Any = _clean(record.get(header))
            if header == "Anul" and re.fullmatch(r"\d{4}", value):
                value = int(value)
            values.append(value)
        sheet.append(values)
        row_number = sheet.max_row
        for cell in sheet[row_number]:
            cell.font = Font(name="Arial", size=10, bold=is_article, color="000000")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_article:
                cell.fill = article_fill
        if not is_article:
            sheet.row_dimensions[row_number].outlineLevel = 1

        for header in ("DOI", "Link"):
            column = WOS_HEADERS.index(header) + 1
            cell = sheet.cell(row_number, column)
            if header == "Link":
                cell.alignment = Alignment(vertical="top", wrap_text=False)
            url = _valid_url(cell.value)
            if header == "DOI" and not url and _normal_doi(cell.value):
                url = f"https://doi.org/{_normal_doi(cell.value)}"
            if url:
                cell.hyperlink = url
                cell.style = "Hyperlink"
                cell.font = Font(
                    name="Arial", size=10, bold=is_article, color="0563C1", underline="single"
                )

    _apply_common_sheet_format(sheet, WOS_HEADERS)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_google(path: Path, source: GoogleData, rows: list[GoogleRow]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Articole și citări Google"
    sheet.append(source.headers)
    article_fill = PatternFill("solid", fgColor="E2F0D9")
    link_columns = {
        index
        for index, header in enumerate(source.headers, start=1)
        if _normal_header(header)
        in {_normal_header("Link"), _normal_header("Link articol"), _normal_header("Article Link")}
    }

    for source_row in rows:
        sheet.append([source_row.values.get(header, "") for header in source.headers])
        row_number = sheet.max_row
        for cell in sheet[row_number]:
            cell.font = Font(
                name="Arial", size=10, bold=source_row.is_article, color="000000"
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if source_row.is_article:
                cell.fill = article_fill
        if not source_row.is_article:
            sheet.row_dimensions[row_number].outlineLevel = 1
        for column in link_columns:
            cell = sheet.cell(row_number, column)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            url = _valid_url(cell.value)
            if url:
                cell.hyperlink = url
                cell.font = Font(
                    name="Arial",
                    size=10,
                    bold=source_row.is_article,
                    color="0563C1",
                    underline="single",
                )

    _apply_common_sheet_format(sheet, source.headers)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _resolve_case_insensitive(base_dir: Path, relative_paths: Iterable[str]) -> Path | None:
    for relative in relative_paths:
        candidate = base_dir / relative
        if candidate.exists():
            return candidate.resolve()
    wanted = {Path(relative).name.casefold() for relative in relative_paths}
    for candidate in base_dir.rglob("*"):
        if (
            candidate.is_file()
            and candidate.name.casefold() in wanted
            and "outputs" not in {part.casefold() for part in candidate.parts}
        ):
            return candidate.resolve()
    return None


def gaseste_fisiere_intrare(base_dir: str | Path = ".") -> dict[str, Path]:
    """Găsește automat cele trei fișiere uzuale din dosarul de lucru."""
    base = Path(base_dir).expanduser().resolve()
    specifications = {
        "articole_google": [
            "Articole_Google.xlsx",
            "Articole_Google.xlsm",
            "Articole_Google.xls",
            "Clasificare/Google_Scholar_Bogdan_Adrian_Enache_clasificat.xlsx",
            "Google_Scholar_Bogdan_Adrian_Enache_clasificat.xlsx",
        ],
        "articole_wos": ["Articole_Wos.xlsx", "Articole_WOS.xlsx", "Articole_Wos.xls", "Articole_WOS.xls"],
        "citari_wos": ["Citari_Wos.xlsx", "Citari_WOS.xlsx", "Citari_Wos.xls", "Citari_WOS.xls"],
    }
    found: dict[str, Path] = {}
    missing: list[str] = []
    for key, candidates in specifications.items():
        result = _resolve_case_insensitive(base, candidates)
        if result is None:
            missing.append(key)
        else:
            found[key] = result
    if missing:
        raise FileNotFoundError(
            "Nu au fost găsite: " + ", ".join(missing) + f" în {base}"
        )
    return found


def construieste_fisiere(
    articole_google: str | Path,
    articole_wos: str | Path,
    citari_wos: str | Path,
    output_wos: str | Path = "Articole_Citari_WoS.xlsx",
    output_google: str | Path = "Articole_Citari_Google.xlsx",
    *,
    prag_fuzzy_citari: float = 0.90,
    verbose: bool = True,
) -> dict[str, Any]:
    """Construiește cele două registre, fără duplicate și fără suprapuneri.

    Articolele WoS sunt scrise bold. Sub fiecare articol sunt incluse numai
    citările sale din Google care au corespondent în Citari_Wos. Restul
    înregistrărilor Google sunt exportate separat, fără a modifica sursa.
    """
    google_path = Path(articole_google).expanduser().resolve()
    articles_path = Path(articole_wos).expanduser().resolve()
    citations_path = Path(citari_wos).expanduser().resolve()
    output_wos_path = Path(output_wos).expanduser().resolve()
    output_google_path = Path(output_google).expanduser().resolve()

    for source in (google_path, articles_path, citations_path):
        if not source.exists():
            raise FileNotFoundError(source)
    if output_wos_path == output_google_path:
        raise ValueError("Cele două fișiere de ieșire trebuie să aibă căi diferite.")
    if output_wos_path in {google_path, articles_path, citations_path} or output_google_path in {
        google_path,
        articles_path,
        citations_path,
    }:
        raise ValueError("Fișierele sursă nu pot fi suprascrise.")

    _log(f"Citesc Articole_Google: {google_path.name}", verbose)
    google = read_google(google_path)
    _log(f"Citesc Articole_WoS: {articles_path.name}", verbose)
    raw_articles = read_table(articles_path)
    _log(f"Citesc Citari_WoS: {citations_path.name}", verbose)
    raw_citations = read_table(citations_path)

    citation_items: list[MatchItem] = []
    for index, row in enumerate(raw_citations):
        record = _canonical_from_wos(row)
        if record["Titlu"] or record["DOI"]:
            citation_items.append(MatchItem(record, {"index": index, "raw": row}))
    citation_matcher = RecordMatcher(citation_items)

    group_items = [
        MatchItem(group.parent.canonical, group)
        for group in google.groups
        if group.parent is not None
    ]
    group_matcher = RecordMatcher(group_items)

    valid_articles: list[dict[str, Any]] = []
    skipped_types: list[dict[str, str]] = []
    reserved_articles = DuplicateTracker()
    for row in raw_articles:
        record = _canonical_from_wos(row)
        source_type = _normalize_type(
            _record_value(row, "Tip", "Type", "Publication Type")
        )
        if source_type not in {"J", "C"}:
            skipped_types.append(
                {"tip": _record_value(row, "Publication Type"), "titlu": record["Titlu"]}
            )
            continue
        record["Tip"] = source_type
        if not record["Titlu"] and not record["DOI"]:
            continue
        valid_articles.append({"record": record, "raw": row})
        reserved_articles.add(record)

    output_rows: list[dict[str, Any]] = []
    output_seen = DuplicateTracker()
    moved_google_rows: set[int] = set()
    used_citation_indices: set[int] = set()
    matched_articles = 0
    articles_without_google_match: list[str] = []
    duplicate_wos_articles = 0
    duplicate_citation_occurrences = 0
    citations_invalid_type = 0

    for number, article in enumerate(valid_articles, start=1):
        record = article["record"]
        if output_seen.contains(record):
            duplicate_wos_articles += 1
            continue
        output_rows.append({"record": record, "is_article": True})
        output_seen.add(record)

        group_item = group_matcher.find(record, fuzzy_threshold=0.94)
        citations_added = 0
        if group_item is None:
            articles_without_google_match.append(record["Titlu"])
        else:
            matched_articles += 1
            group: GoogleGroup = group_item.payload
            if group.parent is not None:
                moved_google_rows.add(group.parent.row_number)
            for google_citation in group.citations:
                citation_item = citation_matcher.find(
                    google_citation.canonical, fuzzy_threshold=prag_fuzzy_citari
                )
                if citation_item is None:
                    continue
                wos_citation = citation_item.record
                merged = _merge_records(wos_citation, google_citation.canonical)
                if merged["Tip"] not in {"J", "C"}:
                    citations_invalid_type += 1
                    continue
                if reserved_articles.contains(merged) or output_seen.contains(merged):
                    duplicate_citation_occurrences += 1
                    continue
                output_rows.append({"record": merged, "is_article": False})
                output_seen.add(merged)
                moved_google_rows.add(google_citation.row_number)
                used_citation_indices.add(citation_item.payload["index"])
                citations_added += 1
        _log(
            f"Articol {number}/{len(valid_articles)}: {record['Titlu'][:75]} "
            f"— {citations_added} citări WoS",
            verbose,
        )

    output_matcher = RecordMatcher(
        MatchItem(entry["record"], None) for entry in output_rows
    )
    google_candidates: list[GoogleRow] = []
    removed_from_google = 0
    for row in google.rows:
        is_in_wos = row.row_number in moved_google_rows or output_matcher.find(
            row.canonical, fuzzy_threshold=0.94
        ) is not None
        if is_in_wos:
            removed_from_google += 1
            continue
        google_candidates.append(row)

    # Dacă aceeași lucrare apare atât ca citare, cât și ca articol bold, păstrăm
    # varianta de articol chiar dacă aceasta se află mai târziu în fișier.
    remaining_article_tracker = DuplicateTracker()
    for row in google_candidates:
        if row.is_article:
            remaining_article_tracker.add(row.canonical)

    remaining_google: list[GoogleRow] = []
    google_seen = DuplicateTracker()
    duplicate_google_rows = 0
    for row in google_candidates:
        if not row.is_article and remaining_article_tracker.contains(row.canonical):
            duplicate_google_rows += 1
            continue
        if google_seen.contains(row.canonical):
            duplicate_google_rows += 1
            continue
        google_seen.add(row.canonical)
        remaining_google.append(row)

    cross_overlap = [
        row.row_number
        for row in remaining_google
        if output_matcher.find(row.canonical, fuzzy_threshold=0.94) is not None
    ]
    if cross_overlap:
        raise RuntimeError(
            "Validarea a găsit suprapuneri între ieșiri la rândurile Google: "
            + ", ".join(map(str, cross_overlap[:10]))
        )
    if any(entry["record"]["Tip"] not in {"J", "C"} for entry in output_rows):
        raise RuntimeError("Ieșirea WoS conține un tip diferit de J/C.")

    _log(f"Scriu {output_wos_path.name}", verbose)
    _write_wos(output_wos_path, output_rows)
    _log(f"Scriu {output_google_path.name}", verbose)
    _write_google(output_google_path, google, remaining_google)

    statistics: dict[str, Any] = {
        "articole_wos_sursa": len(raw_articles),
        "articole_wos_scrise": sum(entry["is_article"] for entry in output_rows),
        "citari_wos_scrise": sum(not entry["is_article"] for entry in output_rows),
        "articole_wos_omise_tip_diferit_de_J_C": len(skipped_types),
        "detalii_tipuri_omise": skipped_types,
        "articole_wos_duplicate_omise": duplicate_wos_articles,
        "articole_wos_gasite_in_google": matched_articles,
        "articole_wos_neasociate_in_google": len(articles_without_google_match),
        "titluri_wos_neasociate": articles_without_google_match,
        "citari_distincte_din_Citari_WOS_folosite": len(used_citation_indices),
        "citari_duplicate_omise": duplicate_citation_occurrences,
        "citari_cu_tip_nevalid_omise": citations_invalid_type,
        "randuri_google_sursa": len(google.rows),
        "randuri_sterse_din_google": removed_from_google,
        "duplicate_interne_google_omise": duplicate_google_rows,
        "randuri_google_ramase": len(remaining_google),
        "suprapuneri_intre_fisiere": 0,
        "fisier_wos": str(output_wos_path),
        "fisier_google": str(output_google_path),
    }
    _log(
        f"FINALIZAT: {statistics['articole_wos_scrise']} articole și "
        f"{statistics['citari_wos_scrise']} citări în WoS; "
        f"{statistics['randuri_google_ramase']} rânduri rămase în Google; "
        "0 suprapuneri.",
        verbose,
    )
    return statistics


def verifica_fisiere_generate(
    fisier_wos: str | Path, fisier_google: str | Path
) -> dict[str, Any]:
    """Verifică antetele, fonturile, tipurile și duplicatele din ieșiri."""
    wos_book = load_workbook(fisier_wos, data_only=False)
    wos_sheet = wos_book.active
    headers = [_clean(cell.value) for cell in wos_sheet[1]]
    if headers != WOS_HEADERS:
        raise AssertionError(f"Antete WoS neașteptate: {headers}")

    wos_records: list[dict[str, str]] = []
    bold_articles = 0
    normal_citations = 0
    for row_number in range(2, wos_sheet.max_row + 1):
        record = {
            header: _clean(wos_sheet.cell(row_number, column).value)
            for column, header in enumerate(headers, start=1)
        }
        if record["Tip"] not in {"J", "C"}:
            raise AssertionError(f"Tip invalid la rândul WoS {row_number}: {record['Tip']}")
        if wos_sheet.cell(row_number, 3).font.bold:
            bold_articles += 1
        else:
            normal_citations += 1
        for cell in wos_sheet[row_number]:
            if cell.font.name != "Arial" or cell.font.sz != 10:
                raise AssertionError(f"Font invalid în WoS la {cell.coordinate}")
        wos_records.append(record)

    duplicate_tracker = DuplicateTracker()
    for record in wos_records:
        if duplicate_tracker.contains(record):
            raise AssertionError(f"Duplicat în ieșirea WoS: {record['Titlu']}")
        duplicate_tracker.add(record)

    google_book = load_workbook(fisier_google, data_only=False)
    google_sheet = google_book.active
    google_headers = [_clean(cell.value) for cell in google_sheet[1]]
    google_records: list[dict[str, str]] = []
    google_tracker = DuplicateTracker()
    for row_number in range(2, google_sheet.max_row + 1):
        source = {
            header: google_sheet.cell(row_number, column).value
            for column, header in enumerate(google_headers, start=1)
        }
        canonical = _canonical_from_google(source)
        if google_tracker.contains(canonical):
            raise AssertionError(f"Duplicat în ieșirea Google: {canonical['Titlu']}")
        google_tracker.add(canonical)
        for cell in google_sheet[row_number]:
            if cell.font.name != "Arial" or cell.font.sz != 10:
                raise AssertionError(f"Font invalid în Google la {cell.coordinate}")
        google_records.append(canonical)

    overlap_matcher = RecordMatcher(MatchItem(record, None) for record in wos_records)
    overlaps = [
        record["Titlu"]
        for record in google_records
        if overlap_matcher.find(record, fuzzy_threshold=0.94) is not None
    ]
    if overlaps:
        raise AssertionError(f"Suprapuneri între ieșiri: {overlaps[:10]}")

    return {
        "randuri_wos": len(wos_records),
        "articole_bold_wos": bold_articles,
        "citari_normale_wos": normal_citations,
        "randuri_google": len(google_records),
        "suprapuneri": 0,
        "font": "Arial 10",
    }


def _main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base-dir", default=".")
    parser.add_argument("--articole-google")
    parser.add_argument("--articole-wos")
    parser.add_argument("--citari-wos")
    parser.add_argument("--output-dir", default=".")
    arguments = parser.parse_args()

    automatic = gaseste_fisiere_intrare(arguments.base_dir)
    output_dir = Path(arguments.output_dir).expanduser().resolve()
    stats = construieste_fisiere(
        arguments.articole_google or automatic["articole_google"],
        arguments.articole_wos or automatic["articole_wos"],
        arguments.citari_wos or automatic["citari_wos"],
        output_dir / "Articole_Citari_WoS.xlsx",
        output_dir / "Articole_Citari_Google.xlsx",
    )
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    _main()


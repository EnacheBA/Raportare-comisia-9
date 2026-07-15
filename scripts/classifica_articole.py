"""Clasifică înregistrările bibliografice ca jurnal (J) sau conferință (C).

Regula solicitată este intenționat simplă: o denumire este conferință dacă
include conference, symposium, proceedings sau un acronim fără spații între
paranteze rotunde. Clasificarea automată trebuie verificată manual.
"""

from __future__ import annotations

import re
import unicodedata
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


KEYWORD_RE = re.compile(r"\b(?:conference|symposium|proceedings)\b", re.IGNORECASE)
ACRONYM_RE = re.compile(r"\(\s*[a-z][a-z0-9&+./-]{1,14}\s*\)", re.IGNORECASE)
VENUE_HEADERS = {
    "revista conferinta",
    "revista conferința",
    "revista / conferinta",
    "revista / conferința",
    "journal conference",
    "source title",
}
TYPE_HEADERS = {"type", "tip", "j c", "j/c"}


def _normal(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text)


def clasifica_denumirea(value: Any) -> str:
    """Returnează C, J sau șir gol."""
    text = str(value or "").strip()
    if not text:
        return ""
    return "C" if KEYWORD_RE.search(text) or ACRONYM_RE.search(text) else "J"


def _gaseste_coloana_antet(worksheet, aliases: set[str]) -> int | None:
    normal_aliases = {_normal(alias) for alias in aliases}
    for column in range(1, worksheet.max_column + 1):
        if _normal(worksheet.cell(1, column).value) in normal_aliases:
            return column
    return None


def _copiaza_stil(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def clasifica_fisier(
    input_path: str | Path,
    output_path: str | Path,
    *,
    sheet_name: str | None = None,
    progress_every: int = 250,
) -> dict[str, Any]:
    """Adaugă/completează coloana Type și salvează într-un fișier nou."""
    source = Path(input_path).expanduser().resolve()
    output = Path(output_path).expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(source)
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Fișierul Google trebuie să fie .xlsx sau .xlsm.")
    if source == output:
        raise ValueError("Fișierul-sursă nu poate fi suprascris.")
    if progress_every < 1:
        raise ValueError("progress_every trebuie să fie cel puțin 1.")

    keep_vba = source.suffix.lower() == ".xlsm"
    workbook = load_workbook(source, data_only=False, keep_vba=keep_vba)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        type_column = _gaseste_coloana_antet(worksheet, TYPE_HEADERS)
        if type_column is None:
            worksheet.insert_cols(1)
            type_column = 1
            worksheet.cell(1, 1).value = "Type"
            if worksheet.max_column >= 2:
                _copiaza_stil(worksheet.cell(1, 2), worksheet.cell(1, 1))

        venue_column = _gaseste_coloana_antet(worksheet, VENUE_HEADERS)
        if venue_column is None:
            raise ValueError(
                "Nu am găsit coloana «Revista / Conferința». "
                f"Antete: {[worksheet.cell(1, c).value for c in range(1, worksheet.max_column + 1)]}"
            )

        counts = Counter({"C": 0, "J": 0, "gol": 0})
        total = max(0, worksheet.max_row - 1)
        for index, row in enumerate(range(2, worksheet.max_row + 1), start=1):
            result = clasifica_denumirea(worksheet.cell(row, venue_column).value)
            target = worksheet.cell(row, type_column)
            target.value = result or None
            reference_column = 2 if type_column == 1 and worksheet.max_column >= 2 else venue_column
            _copiaza_stil(worksheet.cell(row, reference_column), target)
            counts[result if result else "gol"] += 1
            if index % progress_every == 0 or index == total:
                print(f"Clasificare Google: {index}/{total} rânduri")

        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        workbook.close()

    return {
        "fisier": str(output),
        "randuri": total,
        "conferinte_C": counts["C"],
        "jurnale_J": counts["J"],
        "fara_sursa": counts["gol"],
    }


def _self_test() -> None:
    cases = {
        "International Conference on Smart Systems": "C",
        "annual SYMPOSIUM on energy": "C",
        "Proceedings of the IEEE": "C",
        "Advanced Topics in Electrical Engineering (ATEE)": "C",
        "Sensors": "J",
        "Journal of Energy Systems": "J",
        "": "",
        None: "",
    }
    for value, expected in cases.items():
        result = clasifica_denumirea(value)
        if result != expected:
            raise AssertionError((value, expected, result))


if __name__ == "__main__":
    _self_test()
    print("Testele clasificării au trecut.")


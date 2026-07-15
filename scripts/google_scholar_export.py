"""Exportă un profil Google Scholar și lucrările care citează fiecare publicație.

Moduri de acces:
1. ``scholarly`` – acces direct, fără cheie API; Google poate răspunde cu CAPTCHA.
2. ``serpapi`` – acces stabil prin SerpApi; necesită variabila SERPAPI_API_KEY.

Metadatele lipsă (DOI, ISSN/eISSN, volum, număr, pagini) sunt completate, când
există o potrivire suficient de sigură, prin API-ul public Crossref.
"""

from __future__ import annotations

import hashlib
import json
import multiprocessing as mp
import os
import queue as queue_module
import random
import re
import sys
import time
import traceback
import unicodedata
from dataclasses import asdict, dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Callable, Iterable, Optional
from urllib.parse import quote

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


PROFILE_URL = ""
DEFAULT_AUTHOR_ID = ""
HEADERS = [
    "Autori",
    "Titlu",
    "Revista / Conferința",
    "ISSN/eISSN",
    "Anul",
    "Volum",
    "Număr",
    "Pagini",
    "DOI",
    "Link articol",
]
DOI_RE = re.compile(r"10\.\d{4,9}/[-._;()/:A-Z0-9]+", re.IGNORECASE)
YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def _safe_print(value: Any) -> None:
    text = str(value)
    try:
        print(text, flush=True)
    except UnicodeEncodeError:
        encoding = getattr(sys.stdout, "encoding", None) or "utf-8"
        safe = text.encode(encoding, errors="replace").decode(encoding, errors="replace")
        print(safe, flush=True)


@dataclass
class ExportConfig:
    author_id: str = DEFAULT_AUTHOR_ID
    profile_url: str = PROFILE_URL
    output_path: Path = Path("data/intermediate/Articole_Google_neclasificat.xlsx")
    cache_dir: Path = Path(".scholar_cache")
    backend: str = "auto"  # auto | scholarly | serpapi
    serpapi_api_key: str = os.getenv("SERPAPI_API_KEY", "")
    scraperapi_key: str = os.getenv("SCRAPERAPI_KEY", "")
    crossref_mailto: str = os.getenv("CROSSREF_MAILTO", "")
    force_refresh: bool = False
    max_articles: Optional[int] = None
    max_citations_per_article: Optional[int] = None
    request_timeout: int = 60
    scholar_delay_min: float = 2.0
    scholar_delay_max: float = 4.0
    crossref_delay: float = 0.12
    progress_interval: int = 15
    scholar_stall_timeout: int = 180
    checkpoint_every_citations: int = 10

    def __post_init__(self) -> None:
        self.output_path = Path(self.output_path)
        self.cache_dir = Path(self.cache_dir)
        if self.backend not in {"auto", "scholarly", "serpapi"}:
            raise ValueError("backend trebuie să fie: auto, scholarly sau serpapi")
        if self.progress_interval < 1:
            raise ValueError("progress_interval trebuie să fie cel puțin 1 secundă")
        if self.scholar_stall_timeout <= self.progress_interval:
            raise ValueError("scholar_stall_timeout trebuie să fie mai mare decât progress_interval")
        if self.checkpoint_every_citations < 1:
            raise ValueError("checkpoint_every_citations trebuie să fie cel puțin 1")


@dataclass
class BibliographicRecord:
    authors: str = ""
    title: str = ""
    venue: str = ""
    issn: str = ""
    year: str = ""
    volume: str = ""
    number: str = ""
    pages: str = ""
    doi: str = ""
    article_link: str = ""

    def as_excel_row(self) -> list[str]:
        return [
            self.authors,
            self.title,
            self.venue,
            self.issn,
            self.year,
            self.volume,
            self.number,
            self.pages,
            self.doi,
            self.article_link,
        ]


def _plain(value: Any) -> Any:
    """Transformă obiectele scholarly în structuri JSON standard."""
    if isinstance(value, dict):
        return {str(key): _plain(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_plain(item) for item in value]
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _slug(value: str) -> str:
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()
    return digest[:24]


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(_plain(value), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(path)


def _cached(path: Path, force_refresh: bool) -> Any:
    if path.exists() and not force_refresh:
        return _read_json(path)
    return None


class ProgressReporter:
    """Afișează progresul și menține fișiere de stare ușor de inspectat."""

    def __init__(self, config: ExportConfig):
        self.config = config
        self.started = datetime.now().astimezone()
        self.log_path = config.output_path.with_suffix(".log")
        self.status_path = config.output_path.with_suffix(".status.json")
        self.log_path.parent.mkdir(parents=True, exist_ok=True)
        self.state: dict[str, Any] = {
            "state": "starting",
            "stage": "initializare",
            "message": "Pornire export",
            "started_at": self.started.isoformat(timespec="seconds"),
            "updated_at": self.started.isoformat(timespec="seconds"),
            "output": str(config.output_path.resolve()),
            "articles_discovered": 0,
            "articles_cached": 0,
            "citations_cached": 0,
            "complete": False,
        }

    def log(self, message: str, stage: Optional[str] = None, **updates: Any) -> None:
        now = datetime.now().astimezone()
        if stage:
            self.state["stage"] = stage
        self.state.update(updates)
        self.state["message"] = message
        self.state["updated_at"] = now.isoformat(timespec="seconds")
        line = f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] {message}"
        _safe_print(line)
        with self.log_path.open("a", encoding="utf-8") as handle:
            handle.write(line + "\n")
        _write_json(self.status_path, self.state)

    def heartbeat(self, message: str, **updates: Any) -> None:
        self.log(f"ÎNCĂ LUCREAZĂ: {message}", **updates)

    def finish(self, complete: bool, message: str, **updates: Any) -> None:
        self.state["state"] = "complete" if complete else "incomplete"
        self.state["complete"] = complete
        self.log(message, **updates)


def _session(user_agent: str) -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=4,
        read=4,
        connect=4,
        backoff_factor=1.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET"}),
        respect_retry_after_header=True,
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))
    session.headers.update({"User-Agent": user_agent, "Accept": "application/json"})
    return session


def _sleep_between(config: ExportConfig) -> None:
    time.sleep(random.uniform(config.scholar_delay_min, config.scholar_delay_max))


def _normalize_title(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(character for character in value if not unicodedata.combining(character))
    value = value.lower().replace("₄", "4")
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value).split())


def _first_nonempty(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        if isinstance(value, (list, tuple)):
            if value:
                return str(value[0]).strip()
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _authors_text(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        return _first_nonempty(value.get("name"), value.get("family"))
    if isinstance(value, (list, tuple)):
        names = []
        for item in value:
            if isinstance(item, dict):
                name = _first_nonempty(item.get("name"))
                if not name:
                    name = " ".join(
                        part
                        for part in [str(item.get("given", "")).strip(), str(item.get("family", "")).strip()]
                        if part
                    )
                if name:
                    names.append(name)
            elif str(item).strip():
                names.append(str(item).strip())
        return "; ".join(names)
    return str(value).strip()


def _find_doi(value: Any) -> str:
    """Caută DOI în câmpurile furnizate de Scholar/SerpApi."""
    if isinstance(value, dict):
        preferred = ["doi", "DOI", "pub_url", "eprint_url", "link", "citation", "description"]
        for key in preferred:
            if key in value:
                found = _find_doi(value[key])
                if found:
                    return found
        for item in value.values():
            found = _find_doi(item)
            if found:
                return found
        return ""
    if isinstance(value, (list, tuple)):
        for item in value:
            found = _find_doi(item)
            if found:
                return found
        return ""
    match = DOI_RE.search(str(value or ""))
    return match.group(0).rstrip(".,;)]}").lower() if match else ""


def _parse_publication_string(value: str) -> dict[str, str]:
    """Extrage conservator an/volum/număr/pagini din citarea textuală Scholar."""
    result = {"venue": "", "year": "", "volume": "", "number": "", "pages": ""}
    text = " ".join((value or "").replace("–", "-").split())
    if not text:
        return result

    year_matches = list(YEAR_RE.finditer(text))
    if year_matches:
        result["year"] = year_matches[-1].group(0)
        if text.rstrip().endswith(result["year"]):
            text = text[: year_matches[-1].start()].rstrip(" ,")

    pages_match = re.search(r",\s*([A-Za-z]?[0-9]+(?:\s*-\s*[A-Za-z]?[0-9]+)?)$", text)
    if pages_match:
        result["pages"] = pages_match.group(1).replace(" ", "")
        text = text[: pages_match.start()].rstrip(" ,")

    volume_issue = re.search(r"\s+(\d+[A-Za-z]?)\s*\(([^)]+)\)$", text)
    if volume_issue:
        result["volume"] = volume_issue.group(1)
        result["number"] = volume_issue.group(2).strip()
        text = text[: volume_issue.start()].rstrip(" ,")
    else:
        volume_only = re.search(r"\s+(\d+[A-Za-z]?)$", text)
        if volume_only and len(text[: volume_only.start()].strip()) >= 4:
            result["volume"] = volume_only.group(1)
            text = text[: volume_only.start()].rstrip(" ,")

    result["venue"] = text.strip()
    return result


def _summary_parts(raw: dict[str, Any]) -> tuple[str, str]:
    publication_info = raw.get("publication_info") or {}
    summary = str(publication_info.get("summary") or "")
    parts = [part.strip() for part in summary.split(" - ")]
    authors = _authors_text(publication_info.get("authors"))
    if not authors and parts:
        authors = parts[0]
    publication = parts[1] if len(parts) > 1 else ""
    return authors, publication


def raw_to_record(raw: dict[str, Any]) -> BibliographicRecord:
    """Normalizează rezultatele scholarly și SerpApi în aceeași schemă."""
    raw = raw or {}
    bib = raw.get("bib") or {}
    summary_authors, summary_publication = _summary_parts(raw)
    publication = _first_nonempty(
        raw.get("_publication_listing"),
        raw.get("publication"),
        bib.get("citation"),
        summary_publication,
    )
    parsed = _parse_publication_string(publication)

    authors_source = raw.get("authors") or bib.get("author") or summary_authors
    authors = _authors_text(authors_source)
    title = _first_nonempty(raw.get("title"), bib.get("title"))
    venue = _first_nonempty(
        raw.get("journal"),
        raw.get("conference"),
        bib.get("journal"),
        bib.get("venue"),
        bib.get("conference"),
        bib.get("booktitle"),
        parsed["venue"],
    )
    year = _first_nonempty(
        raw.get("year"),
        raw.get("publication_date"),
        bib.get("pub_year"),
        bib.get("year"),
        parsed["year"],
    )
    year_match = YEAR_RE.search(year)
    year = year_match.group(0) if year_match else year

    link = _first_nonempty(
        raw.get("pub_url"),
        raw.get("eprint_url"),
        raw.get("link"),
        bib.get("url"),
    )
    if not link:
        resources = raw.get("resources") or []
        if resources:
            link = _first_nonempty(resources[0].get("link"))

    return BibliographicRecord(
        authors=authors,
        title=title,
        venue=venue,
        year=year,
        volume=_first_nonempty(raw.get("volume"), bib.get("volume"), parsed["volume"]),
        number=_first_nonempty(
            raw.get("issue"), raw.get("number"), bib.get("number"), bib.get("issue"), parsed["number"]
        ),
        pages=_first_nonempty(raw.get("pages"), raw.get("page"), bib.get("pages"), parsed["pages"]),
        doi=_find_doi(raw),
        article_link=link,
    )


class CrossrefEnricher:
    def __init__(self, config: ExportConfig):
        self.config = config
        contact = f"mailto:{config.crossref_mailto}" if config.crossref_mailto else "no-contact"
        self.session = _session(f"GoogleScholarExcelExporter/1.0 ({contact})")
        self.cache_dir = config.cache_dir / "crossref"

    def _get(self, url: str, params: Optional[dict[str, Any]] = None) -> Optional[dict[str, Any]]:
        if params is None:
            params = {}
        if self.config.crossref_mailto:
            params.setdefault("mailto", self.config.crossref_mailto)
        try:
            response = self.session.get(url, params=params, timeout=self.config.request_timeout)
            response.raise_for_status()
            return response.json().get("message")
        except (requests.RequestException, ValueError) as exc:
            _safe_print(f"  Crossref indisponibil pentru această înregistrare: {exc}")
            return None

    @staticmethod
    def _year(message: dict[str, Any]) -> str:
        for key in ("published-print", "published-online", "issued"):
            parts = (message.get(key) or {}).get("date-parts") or []
            if parts and parts[0]:
                return str(parts[0][0])
        return ""

    @staticmethod
    def _authors(message: dict[str, Any]) -> str:
        names = []
        for author in message.get("author") or []:
            name = " ".join(
                part
                for part in [str(author.get("given", "")).strip(), str(author.get("family", "")).strip()]
                if part
            )
            if name:
                names.append(name)
        return "; ".join(names)

    @staticmethod
    def _issn(message: dict[str, Any]) -> str:
        typed = message.get("issn-type") or []
        print_values = [
            item.get("value", "") for item in typed if item.get("type") == "print" and item.get("value")
        ]
        electronic_values = [
            item.get("value", "")
            for item in typed
            if item.get("type") == "electronic" and item.get("value")
        ]
        parts = []
        if print_values:
            parts.append("ISSN " + ", ".join(dict.fromkeys(print_values)))
        if electronic_values:
            parts.append("eISSN " + ", ".join(dict.fromkeys(electronic_values)))
        if parts:
            return "; ".join(parts)
        values = [str(value) for value in message.get("ISSN") or [] if value]
        return ", ".join(dict.fromkeys(values))

    def _best_title_match(
        self, candidates: Iterable[dict[str, Any]], record: BibliographicRecord
    ) -> tuple[Optional[dict[str, Any]], float]:
        normalized = _normalize_title(record.title)
        best: Optional[dict[str, Any]] = None
        best_score = 0.0
        for candidate in candidates:
            candidate_title = _first_nonempty(candidate.get("title"))
            score = SequenceMatcher(None, normalized, _normalize_title(candidate_title)).ratio()
            candidate_year = self._year(candidate)
            if record.year and candidate_year:
                try:
                    delta = abs(int(record.year) - int(candidate_year))
                    if delta > 2:
                        continue
                    if delta == 0:
                        score += 0.03
                except ValueError:
                    pass
            if score > best_score:
                best, best_score = candidate, score
        return best, best_score

    def lookup(self, record: BibliographicRecord) -> Optional[dict[str, Any]]:
        if not record.title and not record.doi:
            return None
        key = _slug(f"{record.doi}|{record.title}|{record.year}|{record.authors}")
        cache_path = self.cache_dir / f"{key}.json"
        cached = _cached(cache_path, self.config.force_refresh)
        if cached is not None:
            return cached.get("message") if cached.get("matched") else None

        message: Optional[dict[str, Any]] = None
        score = 1.0 if record.doi else 0.0
        if record.doi:
            message = self._get(f"https://api.crossref.org/works/{quote(record.doi, safe='')}")
        else:
            params: dict[str, Any] = {"query.title": record.title, "rows": 5}
            first_author = (record.authors.split(";")[0] or "").strip()
            if first_author:
                params["query.author"] = first_author
            payload = self._get("https://api.crossref.org/works", params=params)
            candidates = (payload or {}).get("items") or []
            message, score = self._best_title_match(candidates, record)
            if score < 0.88:
                message = None

        _write_json(
            cache_path,
            {"matched": message is not None, "title_score": round(score, 4), "message": message},
        )
        time.sleep(self.config.crossref_delay)
        return message

    def enrich(self, raw: dict[str, Any]) -> BibliographicRecord:
        record = raw_to_record(raw)
        message = self.lookup(record)
        if not message:
            if record.doi and not record.article_link:
                record.article_link = f"https://doi.org/{record.doi}"
            return record

        crossref_title = _first_nonempty(message.get("title"))
        event = message.get("event") or {}
        if not isinstance(event, dict):
            event = {}
        container = _first_nonempty(message.get("container-title"), event.get("name"))
        crossref_doi = _first_nonempty(message.get("DOI")).lower()
        article_number = _first_nonempty(message.get("article-number"))
        record.authors = record.authors or self._authors(message)
        record.title = record.title or crossref_title
        record.venue = container or record.venue
        record.issn = self._issn(message)
        record.year = record.year or self._year(message)
        record.volume = record.volume or _first_nonempty(message.get("volume"))
        record.number = record.number or _first_nonempty(message.get("issue"))
        record.pages = record.pages or _first_nonempty(message.get("page"), article_number)
        record.doi = record.doi or crossref_doi
        record.article_link = record.article_link or _first_nonempty(message.get("URL"))
        record.pages = record.pages or _first_nonempty(message.get("page"), article_number)
        record.doi = record.doi or crossref_doi
        record.article_link = record.article_link or _first_nonempty(message.get("URL"))
        if record.doi and not record.article_link:
            record.article_link = f"https://doi.org/{record.doi}"
        return record


def _configure_scholarly(scraperapi_key: str):
    """Inițializează scholarly în procesul separat care poate fi oprit la timeout."""
    from scholarly import ProxyGenerator, scholarly

    if scraperapi_key:
        proxy = ProxyGenerator()
        if not proxy.ScraperAPI(scraperapi_key):
            raise RuntimeError("Configurarea ScraperAPI pentru scholarly a eșuat")
        scholarly.use_proxy(proxy)
    return scholarly


def _citation_cache_key(raw: dict[str, Any]) -> str:
    record = raw_to_record(raw)
    return record.doi or f"{_normalize_title(record.title)}|{record.year}"


def _scholarly_collection_worker(config_data: dict[str, Any], events: Any) -> None:
    """Colectează datele într-un proces terminabil și scrie checkpointuri pe disc."""
    try:
        config = ExportConfig(**config_data)
        scholarly = _configure_scholarly(config.scraperapi_key)
        cache_dir = config.cache_dir / "scholarly"
        listing_path = cache_dir / "author_publications.json"
        author = _cached(listing_path, config.force_refresh)

        if author is None:
            events.put(
                {
                    "type": "progress",
                    "stage": "profil Scholar",
                    "message": "Solicit profilul și lista publicațiilor de la Google Scholar...",
                }
            )
            author = scholarly.search_author_id(config.author_id)
            try:
                author = scholarly.fill(author, sections=["basics", "publications"])
            except TypeError:
                author = scholarly.fill(author)
            _write_json(listing_path, author)
            events.put(
                {
                    "type": "listing_saved",
                    "message": "Lista publicațiilor a fost salvată în cache.",
                    "total": len(author.get("publications") or []),
                }
            )
        else:
            events.put(
                {
                    "type": "listing_saved",
                    "message": "Lista publicațiilor a fost încărcată din cache.",
                    "total": len(author.get("publications") or []),
                }
            )

        publications = list(author.get("publications") or [])
        if config.max_articles is not None:
            publications = publications[: config.max_articles]

        for index, publication in enumerate(publications, start=1):
            bib = publication.get("bib") or {}
            title = _first_nonempty(bib.get("title"), publication.get("author_pub_id"))
            key = _slug(_first_nonempty(publication.get("author_pub_id"), title))
            final_path = cache_dir / "groups" / f"{key}.json"
            partial_path = cache_dir / "groups" / f"{key}.partial.json"
            final_cached = _cached(final_path, config.force_refresh)
            if final_cached is not None:
                events.put(
                    {
                        "type": "group_saved",
                        "message": f"Articol {index}/{len(publications)} încărcat din cache: {title}",
                        "index": index,
                        "total": len(publications),
                        "title": title,
                        "citations": len(final_cached.get("citations") or []),
                    }
                )
                continue

            partial = _cached(partial_path, config.force_refresh) or {}
            parent = partial.get("article")
            citations = list(partial.get("citations") or [])
            seen = {_citation_cache_key(item) for item in citations}
            seen.discard("|")

            events.put(
                {
                    "type": "article_started",
                    "message": (
                        f"Articol {index}/{len(publications)}: {title} "
                        f"(citări recuperate anterior: {len(citations)})"
                    ),
                    "index": index,
                    "total": len(publications),
                    "title": title,
                    "citations": len(citations),
                }
            )

            if not parent:
                parent = scholarly.fill(publication)
                _write_json(partial_path, {"article": parent, "citations": citations})

            processed = 0
            if int(parent.get("num_citations") or 0) > 0:
                for citation in scholarly.citedby(parent):
                    processed += 1
                    plain_citation = _plain(citation)
                    citation_key = _citation_cache_key(plain_citation)
                    if citation_key.strip("|") and citation_key not in seen:
                        seen.add(citation_key)
                        citations.append(plain_citation)

                    if processed % config.checkpoint_every_citations == 0:
                        _write_json(partial_path, {"article": parent, "citations": citations})
                        events.put(
                            {
                                "type": "citation_progress",
                                "message": (
                                    f"Articol {index}/{len(publications)}: "
                                    f"{processed} rezultate parcurse, {len(citations)} citări unice salvate."
                                ),
                                "index": index,
                                "total": len(publications),
                                "title": title,
                                "processed": processed,
                                "citations": len(citations),
                            }
                        )

                    limit = config.max_citations_per_article
                    if limit is not None and len(citations) >= limit:
                        citations = citations[:limit]
                        break

            group = {"article": _plain(parent), "citations": citations}
            _write_json(final_path, group)
            partial_path.unlink(missing_ok=True)
            events.put(
                {
                    "type": "group_saved",
                    "message": (
                        f"SALVAT articol {index}/{len(publications)}: {title}; "
                        f"citări: {len(citations)}."
                    ),
                    "index": index,
                    "total": len(publications),
                    "title": title,
                    "citations": len(citations),
                }
            )
            _sleep_between(config)

        events.put({"type": "done", "message": "Colectarea Google Scholar s-a încheiat."})
    except Exception as exc:
        events.put(
            {
                "type": "error",
                "message": f"Colectarea Google Scholar a eșuat: {type(exc).__name__}: {exc}",
                "traceback": traceback.format_exc(limit=8),
            }
        )


class ScholarCollector:
    def __init__(
        self,
        config: ExportConfig,
        reporter: Optional[ProgressReporter] = None,
        on_checkpoint: Optional[Callable[[list[dict[str, Any]]], None]] = None,
    ):
        self.config = config
        self.backend = self._choose_backend()
        self.cache_dir = config.cache_dir / self.backend
        self.reporter = reporter
        self.on_checkpoint = on_checkpoint
        self.complete = True
        self.incomplete_reason = ""

    def _choose_backend(self) -> str:
        if self.config.backend != "auto":
            return self.config.backend
        return "serpapi" if self.config.serpapi_api_key else "scholarly"

    def _emit(self, message: str, stage: str = "Google Scholar", **updates: Any) -> None:
        if self.reporter:
            self.reporter.log(message, stage=stage, **updates)
        else:
            _safe_print(message)

    def collect(self) -> list[dict[str, Any]]:
        self._emit(f"Backend Google Scholar: {self.backend}", stage="selectare backend")
        if self.backend == "serpapi":
            return self._collect_serpapi()
        return self._collect_scholarly()

    def _scholarly_client(self):
        try:
            from scholarly import ProxyGenerator, scholarly
        except ImportError as exc:
            raise RuntimeError(
                "Lipsește pachetul scholarly. Rulați: %pip install scholarly==1.7.11"
            ) from exc

        if self.config.scraperapi_key:
            proxy = ProxyGenerator()
            if not proxy.ScraperAPI(self.config.scraperapi_key):
                raise RuntimeError("Configurarea ScraperAPI pentru scholarly a eșuat")
            scholarly.use_proxy(proxy)
        return scholarly

    def _load_scholarly_checkpoints(self) -> tuple[list[dict[str, Any]], int, int]:
        listing_path = self.cache_dir / "author_publications.json"
        if not listing_path.exists():
            return [], 0, 0
        author = _read_json(listing_path)
        publications = list(author.get("publications") or [])
        if self.config.max_articles is not None:
            publications = publications[: self.config.max_articles]
        groups: list[dict[str, Any]] = []
        completed = 0
        for publication in publications:
            bib = publication.get("bib") or {}
            title = _first_nonempty(bib.get("title"), publication.get("author_pub_id"))
            key = _slug(_first_nonempty(publication.get("author_pub_id"), title))
            final_path = self.cache_dir / "groups" / f"{key}.json"
            partial_path = self.cache_dir / "groups" / f"{key}.partial.json"
            if final_path.exists():
                groups.append(_read_json(final_path))
                completed += 1
            elif partial_path.exists():
                groups.append(_read_json(partial_path))
        return groups, completed, len(publications)

    def _checkpoint_from_scholarly_cache(self) -> None:
        groups, completed, total = self._load_scholarly_checkpoints()
        citations = sum(len(group.get("citations") or []) for group in groups)
        if self.reporter:
            self.reporter.state.update(
                {
                    "articles_discovered": total,
                    "articles_cached": completed,
                    "citations_cached": citations,
                }
            )
        if self.on_checkpoint:
            self.on_checkpoint(groups)

    def _collect_scholarly(self) -> list[dict[str, Any]]:
        try:
            import scholarly  # noqa: F401
        except ImportError as exc:
            raise RuntimeError(
                "Lipsește pachetul scholarly. Rulați: %pip install scholarly==1.7.11"
            ) from exc

        context = mp.get_context("spawn")
        events = context.Queue()
        process = context.Process(
            target=_scholarly_collection_worker,
            args=(_plain(asdict(self.config)), events),
            daemon=True,
        )
        done = False
        last_progress = time.monotonic()
        started = last_progress

        def handle_event(event: dict[str, Any]) -> None:
            nonlocal done, last_progress
            last_progress = time.monotonic()
            event_type = event.get("type", "progress")
            message = event.get("message", event_type)
            updates: dict[str, Any] = {}
            if event.get("total") is not None:
                updates["articles_discovered"] = event.get("total")
            if event.get("index") is not None:
                updates["current_article_index"] = event.get("index")
            if event.get("title"):
                updates["current_article"] = event.get("title")
            if event.get("citations") is not None:
                updates["current_article_citations"] = event.get("citations")

            if event_type == "error":
                self.complete = False
                self.incomplete_reason = message
                if self.reporter:
                    self.reporter.state["last_traceback"] = event.get("traceback", "")
            elif event_type == "done":
                done = True

            self._emit(message, stage="colectare Scholar", **updates)
            if event_type in {"listing_saved", "citation_progress", "group_saved"}:
                self._checkpoint_from_scholarly_cache()

        try:
            process.start()
            while True:
                try:
                    event = events.get(timeout=self.config.progress_interval)
                    handle_event(event)
                except queue_module.Empty:
                    silent_for = int(time.monotonic() - last_progress)
                    elapsed = int(time.monotonic() - started)
                    if self.reporter:
                        self.reporter.heartbeat(
                            f"procesul Scholar răspunde încă; fără progres nou de {silent_for}s "
                            f"(timp total {elapsed}s).",
                            stage="așteptare Scholar",
                            silent_seconds=silent_for,
                            elapsed_seconds=elapsed,
                        )
                    else:
                        _safe_print(
                            f"ÎNCĂ LUCREAZĂ: fără progres nou de {silent_for}s (total {elapsed}s)."
                        )

                if time.monotonic() - last_progress >= self.config.scholar_stall_timeout:
                    self.complete = False
                    self.incomplete_reason = (
                        f"Nicio actualizare de la Google Scholar timp de "
                        f"{self.config.scholar_stall_timeout}s. Procesul a fost oprit; "
                        "checkpointurile au fost păstrate pentru reluare."
                    )
                    self._emit(self.incomplete_reason, stage="timeout Scholar")
                    if process.is_alive():
                        process.terminate()
                        process.join(timeout=10)
                    if process.is_alive():
                        process.kill()
                        process.join(timeout=5)
                    break

                if not process.is_alive():
                    while True:
                        try:
                            handle_event(events.get_nowait())
                        except queue_module.Empty:
                            break
                    break

            process.join(timeout=5)
            if process.exitcode not in (0, None) and not self.incomplete_reason:
                self.complete = False
                self.incomplete_reason = f"Procesul scholarly s-a închis cu codul {process.exitcode}."
                self._emit(self.incomplete_reason, stage="eroare Scholar")
            if not done and not self.incomplete_reason:
                self.complete = False
                self.incomplete_reason = "Colectarea scholarly s-a închis înainte de mesajul final."
                self._emit(self.incomplete_reason, stage="colectare incompletă")
        except Exception as exc:
            self.complete = False
            self.incomplete_reason = f"Nu am putut porni/monitoriza scholarly: {type(exc).__name__}: {exc}"
            self._emit(self.incomplete_reason, stage="eroare Scholar")
        finally:
            if process.is_alive():
                process.terminate()
                process.join(timeout=5)
            events.close()

        groups, completed, total = self._load_scholarly_checkpoints()
        if self.complete and completed < total:
            self.complete = False
            self.incomplete_reason = f"Sunt salvate complet {completed} din {total} articole."
        self._checkpoint_from_scholarly_cache()
        return groups

    @staticmethod
    def _is_serpapi_no_results(message: str) -> bool:
        normalized = " ".join(str(message or "").lower().split())
        return any(
            marker in normalized
            for marker in (
                "hasn't returned any results",
                "has not returned any results",
                "no results for this query",
                "no results found",
            )
        )

    def _serpapi_get(
        self,
        params: dict[str, Any],
        allow_no_results: bool = False,
    ) -> dict[str, Any]:
        if not self.config.serpapi_api_key:
            raise RuntimeError("Pentru backend='serpapi' trebuie setată variabila SERPAPI_API_KEY")
        params = {**params, "api_key": self.config.serpapi_api_key, "output": "json"}
        session = _session("GoogleScholarExcelExporter/1.0")
        response = session.get(
            "https://serpapi.com/search.json",
            params=params,
            timeout=self.config.request_timeout,
        )
        response.raise_for_status()
        payload = response.json()
        if payload.get("error"):
            error = str(payload["error"])
            if allow_no_results and self._is_serpapi_no_results(error):
                return {"_no_results": True, "error": error}
            raise RuntimeError(f"SerpApi: {error}")
        return payload

    @staticmethod
    def _cites_id(article: dict[str, Any], detail: dict[str, Any]) -> str:
        article_cited_by = article.get("cited_by") or {}
        if not isinstance(article_cited_by, dict):
            article_cited_by = {}
        total_citations = detail.get("total_citations") or {}
        if not isinstance(total_citations, dict):
            total_citations = {}
        nested_cited_by = total_citations.get("cited_by") or {}
        if not isinstance(nested_cited_by, dict):
            nested_cited_by = {}
        candidates = [
            article_cited_by.get("cites_id"),
            total_citations.get("cites_id"),
            nested_cited_by.get("cites_id"),
        ]
        return _first_nonempty(*candidates)

    def _serpapi_citations(self, cites_id: str) -> list[dict[str, Any]]:
        if not cites_id:
            return []
        results: list[dict[str, Any]] = []
        seen: set[str] = set()
        start = 0
        while True:
            payload = self._serpapi_get(
                {"engine": "google_scholar", "hl": "en", "cites": cites_id, "num": 20, "start": start},
                allow_no_results=True,
            )
            if payload.get("_no_results"):
                self._emit(
                    "SerpApi nu a returnat citări pentru această pagină; continui cu articolul următor.",
                    stage="fără rezultate SerpApi",
                    current_article_citations=len(results),
                )
                break
            page = payload.get("organic_results") or []
            if not page:
                break
            new_count = 0
            for item in page:
                key = _first_nonempty(item.get("result_id"), item.get("title"))
                if key and key in seen:
                    continue
                if key:
                    seen.add(key)
                item = {**item, "_source": "serpapi_citation"}
                results.append(item)
                new_count += 1
                limit = self.config.max_citations_per_article
                if limit is not None and len(results) >= limit:
                    return results[:limit]
            pagination = payload.get("serpapi_pagination") or {}
            self._emit(
                f"SerpApi citări: {len(results)} rezultate salvate pentru articolul curent.",
                stage="colectare citări SerpApi",
                current_article_citations=len(results),
            )
            if new_count == 0 or not (pagination.get("next") or pagination.get("next_link")):
                break
            start += 20
        return results

    def _collect_serpapi(self) -> list[dict[str, Any]]:
        listing_path = self.cache_dir / "author_publications.json"
        listing = _cached(listing_path, self.config.force_refresh)
        if listing is None:
            self._emit("Solicit lista publicațiilor prin SerpApi...", stage="profil SerpApi")
            listing = self._serpapi_get(
                {
                    "engine": "google_scholar_author",
                    "author_id": self.config.author_id,
                    "hl": "en",
                    "num": 100,
                    "sort": "pubdate",
                }
            )
            _write_json(listing_path, listing)
        else:
            self._emit("Lista publicațiilor SerpApi a fost încărcată din cache.", stage="profil SerpApi")

        articles = list(listing.get("articles") or [])
        if self.config.max_articles is not None:
            articles = articles[: self.config.max_articles]
        if self.reporter:
            self.reporter.state["articles_discovered"] = len(articles)
        groups: list[dict[str, Any]] = []
        if self.on_checkpoint:
            self.on_checkpoint(groups)

        for index, article in enumerate(articles, start=1):
            title = _first_nonempty(article.get("title"), article.get("citation_id"))
            key = _slug(_first_nonempty(article.get("citation_id"), title))
            cache_path = self.cache_dir / "groups" / f"{key}.json"
            cached = _cached(cache_path, self.config.force_refresh)
            if cached is not None:
                groups.append(cached)
                self._emit(
                    f"Articol {index}/{len(articles)} încărcat din cache: {title}",
                    stage="colectare SerpApi",
                    current_article_index=index,
                    current_article=title,
                )
                if self.on_checkpoint:
                    self.on_checkpoint(groups)
                continue

            self._emit(
                f"Articol {index}/{len(articles)} prin SerpApi: {title}",
                stage="colectare SerpApi",
                current_article_index=index,
                current_article=title,
            )
            citation_id = article.get("citation_id")
            if citation_id:
                detail_payload = self._serpapi_get(
                    {
                        "engine": "google_scholar_author",
                        "author_id": self.config.author_id,
                        "citation_id": citation_id,
                        "view_op": "view_citation",
                        "hl": "en",
                    },
                    allow_no_results=True,
                )
            else:
                detail_payload = {
                    "_no_results": True,
                    "error": "Articolul nu are citation_id în lista profilului.",
                }
            if detail_payload.get("_no_results"):
                self._emit(
                    f"FĂRĂ DETALII SerpApi pentru articolul {index}/{len(articles)}; "
                    "păstrez datele din profil și continui.",
                    stage="fără rezultate SerpApi",
                    current_article_index=index,
                    current_article=title,
                )
            detail = detail_payload.get("citation") or {}
            parent = {
                **article,
                "_publication_listing": article.get("publication", ""),
                **detail,
                "_source": "serpapi_author",
            }
            citations = self._serpapi_citations(self._cites_id(article, detail))
            group = {"article": parent, "citations": citations}
            _write_json(cache_path, group)
            groups.append(group)
            if self.on_checkpoint:
                self.on_checkpoint(groups)
        return groups


def _deduplicate_within_group(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique = []
    seen: set[str] = set()
    for raw in records:
        record = raw_to_record(raw)
        key = record.doi or f"{_normalize_title(record.title)}|{record.year}"
        if not key.strip("|") or key in seen:
            continue
        seen.add(key)
        unique.append(raw)
    return unique


def _safe_hyperlink(value: str) -> bool:
    return value.startswith(("https://", "http://")) and len(value) <= 2079


def write_excel(
    groups: list[tuple[BibliographicRecord, list[BibliographicRecord]]],
    output_path: Path,
) -> dict[str, Any]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Articole și citări"
    sheet.append(HEADERS)

    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    article_fill = PatternFill("solid", fgColor="D9EAF7")
    normal_fill = PatternFill("solid", fgColor="FFFFFF")
    article_rows: list[int] = []
    citation_rows: list[int] = []

    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30

    for article, citations in groups:
        sheet.append(article.as_excel_row())
        article_row = sheet.max_row
        article_rows.append(article_row)
        for cell in sheet[article_row]:
            cell.font = Font(name="Arial", size=10, bold=True, color="000000")
            cell.fill = article_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[article_row].height = 42

        citation_start = sheet.max_row + 1
        for citation in citations:
            sheet.append(citation.as_excel_row())
            citation_row = sheet.max_row
            citation_rows.append(citation_row)
            for cell in sheet[citation_row]:
                cell.font = Font(name="Arial", size=10, bold=False, color="000000")
                cell.fill = normal_fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            sheet.cell(citation_row, 2).alignment = Alignment(vertical="top", wrap_text=True, indent=1)
            sheet.row_dimensions[citation_row].height = 36
            sheet.row_dimensions[citation_row].outlineLevel = 1
        if citations:
            sheet.row_dimensions.group(citation_start, sheet.max_row, outline_level=1, hidden=False)

    for row in range(2, sheet.max_row + 1):
        is_article = row in set(article_rows)
        for column in (9, 10):
            cell = sheet.cell(row, column)
            target = str(cell.value or "").strip()
            if column == 9 and target:
                target = f"https://doi.org/{target}"
            if _safe_hyperlink(target):
                cell.hyperlink = target
                cell.font = Font(
                    name="Arial",
                    size=10,
                    bold=is_article,
                    color="0563C1",
                    underline="single",
                )

    widths = [36, 58, 42, 24, 10, 10, 10, 14, 30, 65]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{sheet.max_row}"
    sheet.sheet_properties.outlinePr.summaryBelow = False
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4
    workbook.save(output_path)

    stats = {
        "articles": len(article_rows),
        "citations": len(citation_rows),
        "rows": sheet.max_row,
        "article_rows": article_rows,
        "citation_rows": citation_rows,
    }
    workbook.close()
    return stats


def _existing_excel_stats(output_path: Path) -> dict[str, Any]:
    """Citește statisticile unui checkpoint fără să rescrie fișierul."""
    if not output_path.exists():
        return {
            "articles": 0,
            "citations": 0,
            "rows": 0,
            "article_rows": [],
            "citation_rows": [],
        }

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Articole și citări"]
    article_rows: list[int] = []
    citation_rows: list[int] = []
    for row_number in range(2, sheet.max_row + 1):
        cells = list(sheet[row_number])
        if not any(cell.value not in (None, "") for cell in cells):
            continue
        if any(bool(cell.font.bold) for cell in cells):
            article_rows.append(row_number)
        else:
            citation_rows.append(row_number)

    stats = {
        "articles": len(article_rows),
        "citations": len(citation_rows),
        "rows": sheet.max_row,
        "article_rows": article_rows,
        "citation_rows": citation_rows,
    }
    workbook.close()
    return stats


def validate_excel(output_path: Path, expected: dict[str, Any]) -> dict[str, Any]:
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Articole și citări"]
    problems = []
    for row in expected["article_rows"]:
        for cell in sheet[row]:
            if cell.font.name != "Arial" or cell.font.sz != 10 or not cell.font.bold:
                problems.append(f"Stil articol incorect: {cell.coordinate}")
    for row in expected["citation_rows"]:
        for cell in sheet[row]:
            if cell.font.name != "Arial" or cell.font.sz != 10 or bool(cell.font.bold):
                problems.append(f"Stil citare incorect: {cell.coordinate}")
    if sheet.max_column != len(HEADERS):
        problems.append(f"Număr coloane: {sheet.max_column}, așteptat {len(HEADERS)}")
    workbook.close()
    return {"valid": not problems, "problems": problems[:50]}


def _raw_checkpoint_groups(
    raw_groups: list[dict[str, Any]],
) -> list[tuple[BibliographicRecord, list[BibliographicRecord]]]:
    groups: list[tuple[BibliographicRecord, list[BibliographicRecord]]] = []
    for group in raw_groups:
        article = raw_to_record(group.get("article") or {})
        citations = [
            raw_to_record(raw)
            for raw in _deduplicate_within_group(group.get("citations") or [])
        ]
        groups.append((article, citations))
    return groups


def run_export(config: Optional[ExportConfig] = None) -> dict[str, Any]:
    config = config or ExportConfig()
    config.cache_dir.mkdir(parents=True, exist_ok=True)
    reporter = ProgressReporter(config)
    reporter.state["state"] = "running"
    reporter.log(f"Pornire export. Profil: {config.profile_url}", stage="initializare")

    if not config.output_path.exists():
        write_excel([], config.output_path)
        reporter.log(
            f"Fișierul Excel inițial a fost creat: {config.output_path.name}",
            stage="checkpoint inițial",
        )

    last_raw_checkpoint: list[dict[str, Any]] = []
    last_checkpoint_stats: Optional[dict[str, Any]] = None

    def save_raw_checkpoint(raw_groups: list[dict[str, Any]]) -> None:
        nonlocal last_raw_checkpoint, last_checkpoint_stats
        if not raw_groups and config.output_path.exists():
            reporter.log(
                "Nu există încă grupuri noi; fișierul Excel existent a fost păstrat.",
                stage="checkpoint Excel",
            )
            return
        # Copia este actualizată înainte de scriere, astfel încât grupurile deja
        # colectate să poată fi recuperate chiar dacă checkpointul Excel eșuează.
        last_raw_checkpoint = list(raw_groups)
        checkpoint_groups = _raw_checkpoint_groups(raw_groups)
        stats = write_excel(checkpoint_groups, config.output_path)
        last_checkpoint_stats = stats
        reporter.log(
            f"CHECKPOINT Excel: {stats['articles']} articole și {stats['citations']} citări.",
            stage="checkpoint Excel",
            articles_cached=stats["articles"],
            citations_cached=stats["citations"],
        )

    collector = ScholarCollector(config, reporter=reporter, on_checkpoint=save_raw_checkpoint)
    try:
        raw_groups = collector.collect()
    except Exception as exc:
        collector.complete = False
        collector.incomplete_reason = f"Colectarea s-a oprit: {type(exc).__name__}: {exc}"
        reporter.log(collector.incomplete_reason, stage="eroare colectare")
        raw_groups = list(last_raw_checkpoint)
        if raw_groups:
            reporter.log(
                f"RECUPERARE: continui cu {len(raw_groups)} articole din ultimul checkpoint.",
                stage="recuperare checkpoint",
                articles_cached=len(raw_groups),
                citations_cached=sum(len(group.get("citations") or []) for group in raw_groups),
            )
        else:
            reporter.log(
                "Nu există grupuri recuperabile în checkpoint; fișierul Excel existent nu va fi suprascris.",
                stage="recuperare checkpoint",
            )

    save_raw_checkpoint(raw_groups)
    enricher = CrossrefEnricher(config)
    groups: list[tuple[BibliographicRecord, list[BibliographicRecord]]] = []

    for index, group in enumerate(raw_groups, start=1):
        article_raw = group.get("article") or {}
        reporter.log(
            f"Îmbogățire Crossref articol {index}/{len(raw_groups)}: "
            f"{raw_to_record(article_raw).title}",
            stage="îmbogățire Crossref",
            current_article_index=index,
            current_article=raw_to_record(article_raw).title,
        )
        article = enricher.enrich(article_raw)
        citation_raw = _deduplicate_within_group(group.get("citations") or [])
        citations = []
        for citation_index, raw in enumerate(citation_raw, start=1):
            citations.append(enricher.enrich(raw))
            if citation_index % config.checkpoint_every_citations == 0:
                reporter.log(
                    f"Crossref articol {index}/{len(raw_groups)}: "
                    f"{citation_index}/{len(citation_raw)} citări procesate.",
                    stage="îmbogățire citări",
                    current_article_citations=citation_index,
                )
        groups.append((article, citations))
        remaining_raw = _raw_checkpoint_groups(raw_groups[index:])
        checkpoint_stats = write_excel(groups + remaining_raw, config.output_path)
        reporter.log(
            f"CHECKPOINT îmbogățit: {checkpoint_stats['articles']} articole și "
            f"{checkpoint_stats['citations']} citări.",
            stage="checkpoint Excel îmbogățit",
            articles_enriched=checkpoint_stats["articles"],
            citations_enriched=checkpoint_stats["citations"],
        )

    if groups:
        export_stats = write_excel(groups, config.output_path)
    elif last_checkpoint_stats is not None:
        export_stats = last_checkpoint_stats
        reporter.log(
            "Exportul final nu conține grupuri noi; păstrez ultimul checkpoint Excel.",
            stage="protejare checkpoint",
        )
    else:
        export_stats = _existing_excel_stats(config.output_path)
        reporter.log(
            "Exportul final nu conține grupuri noi; fișierul Excel existent nu a fost rescris.",
            stage="protejare checkpoint",
        )
    validation = validate_excel(config.output_path, export_stats)
    complete = collector.complete and export_stats["articles"] > 0 and validation["valid"]
    incomplete_reason = collector.incomplete_reason
    if not validation["valid"]:
        incomplete_reason = f"Validarea Excel a eșuat: {validation['problems'][:5]}"
    if not raw_groups and not incomplete_reason:
        complete = False
        incomplete_reason = (
            "Nu a fost colectat niciun articol nou. Fișierul Excel existent a fost păstrat; "
            "rulați din nou pentru reluare."
        )
    result = {
        "output": str(config.output_path.resolve()),
        "log": str(reporter.log_path.resolve()),
        "status": str(reporter.status_path.resolve()),
        "backend": collector.backend,
        "articles": export_stats["articles"],
        "citations": export_stats["citations"],
        "rows": export_stats["rows"],
        "complete": complete,
        "incomplete_reason": incomplete_reason,
        "validation": validation,
    }
    summary_path = config.output_path.with_suffix(".summary.json")
    config_summary = _plain(asdict(config))
    config_summary["serpapi_api_key"] = bool(config.serpapi_api_key)
    config_summary["scraperapi_key"] = bool(config.scraperapi_key)
    _write_json(summary_path, {**result, "config": config_summary})
    if complete:
        reporter.finish(
            True,
            f"FINALIZAT: {export_stats['articles']} articole și "
            f"{export_stats['citations']} citări exportate.",
            articles_cached=export_stats["articles"],
            citations_cached=export_stats["citations"],
        )
    else:
        reporter.finish(
            False,
            "EXPORT PARȚIAL. Datele și checkpointurile au fost păstrate. "
            f"Reluați aceeași celulă. Motiv: {incomplete_reason}",
            articles_cached=export_stats["articles"],
            citations_cached=export_stats["citations"],
            incomplete_reason=incomplete_reason,
        )
    _safe_print(json.dumps(result, ensure_ascii=False, indent=2))
    if not validation["valid"]:
        raise RuntimeError(f"Validarea Excel a eșuat: {validation['problems'][:5]}")
    return result


if __name__ == "__main__":
    run_export()

"""Completează șablonul CNATDCU din registrele Articole_Citari WoS/Google.

Scriptul păstrează fișierul sursă nemodificat, copiază stilurile și formulele
rândurilor-model și extinde formulele de total pentru numărul real de intrări.

Exemplu de rulare:
    python completeaza_criterii_cnatdcu.py

Sau cu toate căile precizate explicit:
    python completeaza_criterii_cnatdcu.py --template Criterii_CNATDCU_2025_Enache.xlsx \
        --wos Articole_Citari_WoS.xlsx --google Articole_Citari_Google.xlsx \
        --output Criterii_CNATDCU_2025_Enache_completat.xlsx
"""

from __future__ import annotations

import argparse
import math
import re
import unicodedata
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


SHEET_REPORT = "PROF"
SHEET_SUMMARY = "SINTEZA"
GREEN_FILL = PatternFill(fill_type="solid", fgColor="92D050")
NO_FILL = PatternFill(fill_type=None)
GOOGLE_SUFFIX = "Scopus"

# Coloana A este o clasificare tehnică folosită pentru rutare. Șablonul cere
# bibliografia începând cu autorii, de aceea valoarea J/C nu este inclusă în text.
INCLUDE_CLASSIFICATION_COLUMN_IN_TEXT = False


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Da" if value else "Nu"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def normalized(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value))
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def normalize_type(value: Any) -> str:
    text = clean_text(value).upper()
    return text if text in {"J", "C"} else ""


def is_url_header(header: Any) -> bool:
    marker = normalized(header)
    return "link" in marker or "hyperlink" in marker


def source_cell_value(cell, header: Any) -> Any:
    if is_url_header(header) and cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    value = cell.value
    if isinstance(value, str) and value.upper().startswith("=HYPERLINK("):
        match = re.search(r'HYPERLINK\(\s*"([^"]+)"', value, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return value


@dataclass
class SourceRow:
    row_number: int
    values: list[Any]
    headers: list[str]
    is_article: bool
    publication_type: str

    def value_by_header(self, *aliases: str) -> str:
        wanted = {normalized(alias) for alias in aliases}
        for header, value in zip(self.headers, self.values):
            if normalized(header) in wanted:
                return clean_text(value)
        return ""

    def concatenated(self, suffix: str | None = None) -> str:
        start = 0 if INCLUDE_CLASSIFICATION_COLUMN_IN_TEXT else 1
        parts = [clean_text(value) for value in self.values[start:] if clean_text(value)]
        if suffix:
            parts = [part for part in parts if normalized(part) != normalized(suffix)]
            parts.append(suffix)
        return ", ".join(parts)

    def author_count(self) -> int:
        authors = self.value_by_header("Autori", "Authors", "Author Full Names")
        return count_authors(authors)


@dataclass
class SourceGroup:
    article: SourceRow
    citations: list[SourceRow]


@dataclass
class SourceData:
    path: Path
    rows: list[SourceRow]
    groups: list[SourceGroup]
    orphan_citations: list[SourceRow]


def count_authors(authors: Any) -> int:
    text = clean_text(authors)
    if not text:
        return 1
    semicolon_parts = [part.strip() for part in text.split(";") if part.strip()]
    if len(semicolon_parts) > 1:
        return len(semicolon_parts)
    and_parts = [
        part.strip()
        for part in re.split(r"\s+(?:and|&|și|si)\s+", text, flags=re.IGNORECASE)
        if part.strip()
    ]
    if len(and_parts) > 1:
        return len(and_parts)
    comma_parts = [part.strip() for part in text.split(",") if part.strip()]
    if len(comma_parts) <= 1:
        return 1
    # Format WoS pentru un singur autor: "Nume, AB".
    if len(comma_parts) == 2 and re.fullmatch(r"[A-ZĂÂÎȘȚ.\- ]{1,8}", comma_parts[1]):
        return 1
    # În registrul Google, numele complete sunt frecvent separate prin virgulă.
    return len(comma_parts)


def read_source(path: str | Path) -> SourceData:
    source = Path(path).expanduser().resolve()
    workbook = load_workbook(source, data_only=False, read_only=False)
    sheet = workbook.active
    headers = [clean_text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    title_column = next(
        (
            index
            for index, header in enumerate(headers, start=1)
            if normalized(header) in {normalized("Titlu"), normalized("Article Title"), normalized("Title")}
        ),
        None,
    )
    if title_column is None:
        raise ValueError(f"Nu am găsit coloana de titlu în {source.name}.")

    rows: list[SourceRow] = []
    groups: list[SourceGroup] = []
    orphan_citations: list[SourceRow] = []
    current: SourceGroup | None = None
    for row_number in range(2, sheet.max_row + 1):
        values = [
            source_cell_value(sheet.cell(row_number, column), headers[column - 1])
            for column in range(1, sheet.max_column + 1)
        ]
        if not any(clean_text(value) for value in values):
            continue
        is_article = bool(sheet.cell(row_number, title_column).font.bold)
        row = SourceRow(
            row_number=row_number,
            values=values,
            headers=headers,
            is_article=is_article,
            publication_type=normalize_type(values[0] if values else ""),
        )
        rows.append(row)
        if is_article:
            current = SourceGroup(article=row, citations=[])
            groups.append(current)
        elif current is None:
            orphan_citations.append(row)
        else:
            current.citations.append(row)
    return SourceData(source, rows, groups, orphan_citations)


@dataclass
class TemplateCell:
    value: Any
    style: Any
    hyperlink: Any
    comment: Any


@dataclass
class RowTemplate:
    origin_row: int
    cells: list[TemplateCell]
    height: float | None
    hidden: bool
    outline_level: int
    collapsed: bool


def capture_row_template(sheet, row: int, max_column: int) -> RowTemplate:
    cells: list[TemplateCell] = []
    for column in range(1, max_column + 1):
        cell = sheet.cell(row, column)
        cells.append(
            TemplateCell(
                value=cell.value,
                style=copy(cell._style),
                hyperlink=copy(cell.hyperlink),
                comment=copy(cell.comment),
            )
        )
    dimension = sheet.row_dimensions[row]
    return RowTemplate(
        origin_row=row,
        cells=cells,
        height=dimension.height,
        hidden=bool(dimension.hidden),
        outline_level=int(dimension.outlineLevel or 0),
        collapsed=bool(dimension.collapsed),
    )


def apply_row_template(sheet, target_row: int, template: RowTemplate) -> None:
    for column, template_cell in enumerate(template.cells, start=1):
        cell = sheet.cell(target_row, column)
        cell._style = copy(template_cell.style)
        cell.hyperlink = copy(template_cell.hyperlink)
        cell.comment = copy(template_cell.comment)
        value = template_cell.value
        if isinstance(value, str) and value.startswith("="):
            origin = f"{get_column_letter(column)}{template.origin_row}"
            destination = f"{get_column_letter(column)}{target_row}"
            value = Translator(value, origin=origin).translate_formula(destination)
        cell.value = value
    dimension = sheet.row_dimensions[target_row]
    dimension.height = template.height
    dimension.hidden = template.hidden
    dimension.outlineLevel = template.outline_level
    dimension.collapsed = template.collapsed


@dataclass
class OutputEntry:
    text: str
    role: str  # article | cited | citing
    author_count: int = 1


@dataclass
class Replacement:
    key: str
    header_row: int
    data_start: int
    data_end: int
    total_row: int
    entries: list[OutputEntry]
    template_article: RowTemplate | None = None
    template_cited: RowTemplate | None = None
    template_citing: RowTemplate | None = None

    @property
    def old_count(self) -> int:
        return self.data_end - self.data_start + 1

    @property
    def delta(self) -> int:
        return len(self.entries) - self.old_count


@dataclass
class FormulaSnapshot:
    sheet_name: str
    row: int
    column: int
    formula: str


def capture_formulas(workbook) -> list[FormulaSnapshot]:
    result: list[FormulaSnapshot] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    result.append(FormulaSnapshot(sheet.title, cell.row, cell.column, cell.value))
    return result


def find_row(sheet, label: str, *, exact: bool = True) -> int:
    wanted = normalized(label)
    matches: list[int] = []
    for row in range(1, sheet.max_row + 1):
        current = normalized(sheet.cell(row, 1).value)
        if (exact and current == wanted) or (not exact and current.startswith(wanted)):
            matches.append(row)
    if len(matches) != 1:
        raise ValueError(f"Eticheta {label!r} are {len(matches)} potriviri: {matches}")
    return matches[0]


def map_original_row(row: int, replacements: Iterable[Replacement]) -> int:
    return row + sum(replacement.delta for replacement in replacements if replacement.total_row <= row)


def estimate_height(text: str) -> float:
    lines = max(1, math.ceil(max(len(text), 1) / 165))
    return min(14.7 * lines, 58.8)


def set_data_font(sheet, row: int, *, bold: bool) -> None:
    for column in range(1, 5):
        cell = sheet.cell(row, column)
        color = "000000"
        cell.font = Font(name="Arial", size=10, bold=bold, color=color)
        cell.alignment = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical or "center",
            text_rotation=cell.alignment.text_rotation,
            wrap_text=True,
            shrink_to_fit=cell.alignment.shrink_to_fit,
            indent=cell.alignment.indent,
        )


def apply_entry(sheet, row: int, entry: OutputEntry, replacement: Replacement) -> None:
    if entry.role == "article":
        if replacement.template_article is None:
            raise ValueError(f"Lipsește rândul-model pentru {replacement.key}")
        apply_row_template(sheet, row, replacement.template_article)
        sheet.cell(row, 1).value = entry.text
        sheet.cell(row, 2).value = None
        sheet.cell(row, 3).value = max(1, entry.author_count)
        set_data_font(sheet, row, bold=False)
        sheet.row_dimensions[row].height = estimate_height(entry.text)
        return

    template = replacement.template_cited if entry.role == "cited" else replacement.template_citing
    if template is None:
        raise ValueError(f"Lipsește rândul-model de citări pentru {replacement.key}")
    apply_row_template(sheet, row, template)
    prefix = "Articol citat: " if entry.role == "cited" else "Articol care citează: "
    sheet.cell(row, 1).value = prefix + entry.text
    sheet.cell(row, 2).value = None
    sheet.cell(row, 3).value = None
    sheet.cell(row, 4).value = None if entry.role == "cited" else 2
    for column in range(1, 5):
        sheet.cell(row, column).fill = copy(GREEN_FILL if entry.role == "cited" else NO_FILL)
    set_data_font(sheet, row, bold=entry.role == "cited")
    sheet.row_dimensions[row].height = estimate_height(prefix + entry.text)


def set_sum_formula(sheet, total_row: int, start_row: int, end_row: int) -> None:
    sheet.cell(total_row, 4).value = f"=SUM(D{start_row}:D{end_row})" if end_row >= start_row else "=0"


def rebuild_merges(sheet, original_merges: list[str], replacements: list[Replacement]) -> None:
    replaced_blocks = [(item.data_start, item.data_end) for item in replacements]
    for merged in original_merges:
        min_col, min_row, max_col, max_row = range_boundaries(merged)
        if any(start <= min_row and max_row <= end for start, end in replaced_blocks):
            continue
        mapped_start = map_original_row(min_row, replacements)
        mapped_end = map_original_row(max_row, replacements)
        sheet.merge_cells(
            start_row=mapped_start,
            start_column=min_col,
            end_row=mapped_end,
            end_column=max_col,
        )

    for replacement in replacements:
        row = map_original_row(replacement.data_start, replacements)
        for entry in replacement.entries:
            if entry.role == "cited":
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            elif entry.role == "citing":
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            row += 1


def restore_row_dimensions(sheet, original_dimensions: dict[int, dict[str, Any]], replacements: list[Replacement]) -> None:
    for row in range(1, sheet.max_row + 1):
        dimension = sheet.row_dimensions[row]
        dimension.height = None
        dimension.hidden = False
        dimension.outlineLevel = 0
        dimension.collapsed = False
    replaced_blocks = [(item.data_start, item.data_end) for item in replacements]
    for original_row, attributes in original_dimensions.items():
        if any(start <= original_row <= end for start, end in replaced_blocks):
            continue
        mapped = map_original_row(original_row, replacements)
        dimension = sheet.row_dimensions[mapped]
        dimension.height = attributes["height"]
        dimension.hidden = attributes["hidden"]
        dimension.outlineLevel = attributes["outline"]
        dimension.collapsed = attributes["collapsed"]


def restore_unaffected_formulas(
    workbook,
    snapshots: list[FormulaSnapshot],
    replacements: list[Replacement],
) -> None:
    explicit_prof_rows = {
        33,
        37,
        42,
        47,
        51,
        52,
        91,
        102,
        110,
        111,
        112,
        113,
    }
    replaced_blocks = [(item.data_start, item.data_end) for item in replacements]
    explicit_summary_cells = {(6, 5), (7, 5), (8, 5), (10, 5)}

    for snapshot in snapshots:
        if snapshot.sheet_name == SHEET_REPORT:
            if snapshot.row in explicit_prof_rows:
                continue
            if any(start <= snapshot.row <= end for start, end in replaced_blocks):
                continue
            target_row = map_original_row(snapshot.row, replacements)
            target_column = snapshot.column
        else:
            if snapshot.sheet_name == SHEET_SUMMARY and (snapshot.row, snapshot.column) in explicit_summary_cells:
                continue
            target_row = snapshot.row
            target_column = snapshot.column

        origin = f"{get_column_letter(snapshot.column)}{snapshot.row}"
        destination = f"{get_column_letter(target_column)}{target_row}"
        try:
            formula = Translator(snapshot.formula, origin=origin).translate_formula(destination)
        except Exception:
            formula = snapshot.formula
        workbook[snapshot.sheet_name].cell(target_row, target_column).value = formula


def repair_totals(workbook, replacements: list[Replacement]) -> None:
    sheet = workbook[SHEET_REPORT]
    by_key = {replacement.key: replacement for replacement in replacements}

    for key in ("2.1.a", "2.1.b", "2.2.a", "2.2.b", "3.1.a", "3.1.b"):
        replacement = by_key[key]
        total_row = map_original_row(replacement.total_row, replacements)
        start_row = map_original_row(replacement.data_start, replacements)
        end_row = start_row + len(replacement.entries) - 1
        set_sum_formula(sheet, total_row, start_row, end_row)

    total_21a = find_row(sheet, "Total 2.1.a.")
    total_21b = find_row(sheet, "Total 2.1.b.")
    total_21c = find_row(sheet, "Total 2.1.c.")
    total_21 = find_row(sheet, "Total 2.1.")
    sheet.cell(total_21, 4).value = f"=D{total_21c}+D{total_21b}+D{total_21a}"

    total_22a = find_row(sheet, "Total 2.2.a.")
    total_22b = find_row(sheet, "Total 2.2.b.")
    total_22 = find_row(sheet, "Total 2.2.")
    sheet.cell(total_22, 4).value = f"=D{total_22a}+D{total_22b}"

    total_23 = find_row(sheet, "Total 2.3.")
    total_a2 = find_row(sheet, "TOTAL A2. Activitatea de cercetare")
    sheet.cell(total_a2, 4).value = f"=D{total_23}+D{total_22}+D{total_21}"

    total_31a = find_row(sheet, "Total 3.1.a.")
    total_31b = find_row(sheet, "Total 3.1.b.")
    total_31 = find_row(sheet, "Total 3.1.")
    total_a3 = find_row(sheet, "TOTAL A3. Recunoașterea si impactul activității")
    sheet.cell(total_31, 4).value = f"=D{total_31b}+D{total_31a}"
    sheet.cell(total_a3, 4).value = f"=D{total_31}"

    total_a1 = find_row(sheet, "TOTAL A1. Activitatea didactică și profesională")
    total_a = find_row(sheet, "TOTAL A = A1 +A2 + A3")
    sheet.cell(total_a, 4).value = f"=D{total_a3}+D{total_a2}+D{total_a1}"

    summary = workbook[SHEET_SUMMARY]
    summary["E6"] = f"='PROF'!D{total_21}"
    summary["E7"] = f"='PROF'!D{total_22}"
    summary["E8"] = f"='PROF'!D{total_23}"
    summary["E10"] = f"='PROF'!D{total_31}"


def build_replacements(sheet, wos: SourceData, google: SourceData) -> list[Replacement]:
    wos_articles_j = [row for row in wos.rows if row.is_article and row.publication_type == "J"]
    wos_articles_c = [row for row in wos.rows if row.is_article and row.publication_type == "C"]
    google_articles_j = [row for row in google.rows if row.is_article and row.publication_type == "J"]
    google_articles_c = [row for row in google.rows if row.is_article and row.publication_type == "C"]

    wos_groups = [group for group in wos.groups if group.citations]
    google_groups = [group for group in google.groups if group.citations]

    article_entries = lambda rows, suffix=None: [
        OutputEntry(row.concatenated(suffix), "article", row.author_count()) for row in rows
    ]
    citation_entries = lambda groups, suffix=None: [
        entry
        for group in groups
        for entry in (
            [OutputEntry(group.article.concatenated(suffix), "cited")]
            + [OutputEntry(row.concatenated(suffix), "citing") for row in group.citations]
        )
    ]

    return [
        Replacement(
            "2.1.a", 30, 31, 32, 33, article_entries(wos_articles_j),
            template_article=capture_row_template(sheet, 31, sheet.max_column),
        ),
        Replacement(
            "2.1.b", 34, 35, 36, 37, article_entries(wos_articles_c),
            template_article=capture_row_template(sheet, 35, sheet.max_column),
        ),
        Replacement(
            "2.2.a", 44, 45, 46, 47, article_entries(google_articles_j, GOOGLE_SUFFIX),
            template_article=capture_row_template(sheet, 45, sheet.max_column),
        ),
        Replacement(
            "2.2.b", 48, 49, 50, 51, article_entries(google_articles_c, GOOGLE_SUFFIX),
            template_article=capture_row_template(sheet, 49, sheet.max_column),
        ),
        Replacement(
            "3.1.a", 95, 96, 101, 102, citation_entries(wos_groups),
            template_cited=capture_row_template(sheet, 96, sheet.max_column),
            template_citing=capture_row_template(sheet, 97, sheet.max_column),
        ),
        Replacement(
            "3.1.b", 103, 104, 109, 110, citation_entries(google_groups, GOOGLE_SUFFIX),
            template_cited=capture_row_template(sheet, 104, sheet.max_column),
            template_citing=capture_row_template(sheet, 105, sheet.max_column),
        ),
    ]


def validate_completed_workbook(workbook, replacements: list[Replacement]) -> None:
    """Oprește scriptul dacă structura, stilurile sau formulele-cheie sunt invalide."""
    sheet = workbook[SHEET_REPORT]
    merged = {str(item) for item in sheet.merged_cells.ranges}
    for replacement in replacements:
        start = map_original_row(replacement.data_start, replacements)
        end = start + len(replacement.entries) - 1
        total = map_original_row(replacement.total_row, replacements)
        expected_total = f"=SUM(D{start}:D{end})" if end >= start else "=0"
        if sheet.cell(total, 4).value != expected_total:
            raise AssertionError(f"Formula totalului {replacement.key} este incorectă.")
        for offset, entry in enumerate(replacement.entries):
            row = start + offset
            cell = sheet.cell(row, 1)
            if cell.font.name != "Arial" or cell.font.sz != 10:
                raise AssertionError(f"Font invalid la {cell.coordinate}.")
            if replacement.key in {"2.2.a", "2.2.b", "3.1.b"} and not clean_text(cell.value).endswith(GOOGLE_SUFFIX):
                raise AssertionError(f"Lipsește sufixul Google Scholar la {cell.coordinate}.")
            if entry.role == "cited":
                if cell.fill.fgColor.rgb not in {"0092D050", "FF92D050"}:
                    raise AssertionError(f"Fundalul verde lipsește la {cell.coordinate}.")
                if f"A{row}:D{row}" not in merged:
                    raise AssertionError(f"Îmbinare invalidă pentru articolul citat la rândul {row}.")
            elif entry.role == "citing":
                if sheet.cell(row, 4).value != 2:
                    raise AssertionError(f"Punctaj invalid pentru citare la rândul {row}.")
                if f"A{row}:C{row}" not in merged:
                    raise AssertionError(f"Îmbinare invalidă pentru citare la rândul {row}.")

    formulas = [
        cell.value
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    if any("#REF!" in formula.upper() for formula in formulas):
        raise AssertionError("A fost detectată o referință #REF! în formule.")


def complete_workbook(
    template_path: str | Path,
    wos_path: str | Path,
    google_path: str | Path,
    output_path: str | Path,
) -> dict[str, Any]:
    template = Path(template_path).expanduser().resolve()
    output = Path(output_path).expanduser().resolve()
    if output == template:
        raise ValueError("Fișierul șablon nu poate fi suprascris. Alegeți alt fișier de ieșire.")
    for source in (template, Path(wos_path), Path(google_path)):
        if not Path(source).expanduser().exists():
            raise FileNotFoundError(source)

    print(f"Citesc sursa WoS: {Path(wos_path).name}")
    wos = read_source(wos_path)
    print(f"Citesc sursa Google: {Path(google_path).name}")
    google = read_source(google_path)
    print(f"Citesc șablonul: {template.name}")
    workbook = load_workbook(template, data_only=False, rich_text=True)
    if SHEET_REPORT not in workbook.sheetnames or SHEET_SUMMARY not in workbook.sheetnames:
        raise ValueError("Șablonul trebuie să conțină foile PROF și SINTEZA.")
    sheet = workbook[SHEET_REPORT]

    formula_snapshots = capture_formulas(workbook)
    original_merges = [str(merged) for merged in sheet.merged_cells.ranges]
    original_dimensions = {
        row: {
            "height": sheet.row_dimensions[row].height,
            "hidden": bool(sheet.row_dimensions[row].hidden),
            "outline": int(sheet.row_dimensions[row].outlineLevel or 0),
            "collapsed": bool(sheet.row_dimensions[row].collapsed),
        }
        for row in range(1, sheet.max_row + 1)
    }
    replacements = build_replacements(sheet, wos, google)

    for merged in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged))

    # Procesarea de jos în sus păstrează coordonatele originale ale blocurilor
    # care urmează să fie extinse.
    for replacement in sorted(replacements, key=lambda item: item.header_row, reverse=True):
        if replacement.delta > 0:
            sheet.insert_rows(replacement.total_row, replacement.delta)
        elif replacement.delta < 0:
            sheet.delete_rows(
                replacement.data_start + len(replacement.entries),
                -replacement.delta,
            )
        print(f"{replacement.key}: {len(replacement.entries)} rânduri introduse")

    restore_row_dimensions(sheet, original_dimensions, replacements)
    restore_unaffected_formulas(workbook, formula_snapshots, replacements)
    # Rândurile-model se aplică după toate inserările, astfel formulele lor sunt
    # traduse direct către coordonatele finale.
    for replacement in replacements:
        final_start = map_original_row(replacement.data_start, replacements)
        for offset, entry in enumerate(replacement.entries):
            apply_entry(sheet, final_start + offset, entry, replacement)
    rebuild_merges(sheet, original_merges, replacements)
    repair_totals(workbook, replacements)
    validate_completed_workbook(workbook, replacements)
    print("Validare internă: OK")

    if hasattr(workbook, "calculation"):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)

    statistics = {
        "articole_wos_j": sum(row.is_article and row.publication_type == "J" for row in wos.rows),
        "articole_wos_c": sum(row.is_article and row.publication_type == "C" for row in wos.rows),
        "articole_google_j": sum(row.is_article and row.publication_type == "J" for row in google.rows),
        "articole_google_c": sum(row.is_article and row.publication_type == "C" for row in google.rows),
        "grupuri_citari_wos": sum(bool(group.citations) for group in wos.groups),
        "citari_wos": sum(len(group.citations) for group in wos.groups if group.citations),
        "grupuri_citari_google": sum(bool(group.citations) for group in google.groups),
        "citari_google": sum(len(group.citations) for group in google.groups if group.citations),
        "citari_google_fara_parinte_bold_omise": len(google.orphan_citations),
        "randuri_finale_prof": workbook[SHEET_REPORT].max_row,
        "validare": "OK",
        "fisier": str(output),
    }
    return statistics


def find_default_file(base: Path, names: Iterable[str]) -> Path:
    for name in names:
        candidate = base / name
        if candidate.exists():
            return candidate.resolve()
    wanted = {name.casefold() for name in names}
    for candidate in base.rglob("*.xlsx"):
        if candidate.name.casefold() in wanted and "work" not in {part.casefold() for part in candidate.parts}:
            return candidate.resolve()
    raise FileNotFoundError("Nu am găsit niciunul dintre fișierele: " + ", ".join(names))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base-dir", default=".", help="Dosarul de lucru")
    parser.add_argument("--template", help="Șablonul Criterii_CNATDCU_2025_Enache.xlsx")
    parser.add_argument("--wos", help="Fișierul Articole_Citari_WoS.xlsx")
    parser.add_argument("--google", help="Fișierul Articole_Citari_Google.xlsx")
    parser.add_argument("--output", help="Fișierul rezultat")
    args = parser.parse_args()

    base = Path(args.base_dir).expanduser().resolve()
    template = Path(args.template).expanduser().resolve() if args.template else find_default_file(
        base, ["Criterii_CNATDCU_2025_Enache.xlsx"]
    )
    wos = Path(args.wos).expanduser().resolve() if args.wos else find_default_file(
        base, ["Articole_Citari_WoS.xlsx", "Articole_Citari_Wos.xlsx"]
    )
    google = Path(args.google).expanduser().resolve() if args.google else find_default_file(
        base, ["Articole_Citari_Google.xlsx"]
    )
    output = Path(args.output).expanduser().resolve() if args.output else base / "Criterii_CNATDCU_2025_Enache_completat.xlsx"

    stats = complete_workbook(template, wos, google, output)
    print("FINALIZAT")
    for key, value in stats.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()


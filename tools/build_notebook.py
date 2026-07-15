"""Generează notebookul principal dintr-o sursă Python versionabilă."""

from __future__ import annotations

import json
from pathlib import Path
from textwrap import dedent


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Raportare_OMEC_3019_2025.ipynb"


def _source(text: str) -> list[str]:
    value = dedent(text).strip("\n") + "\n"
    return value.splitlines(keepends=True)


def markdown(text: str) -> dict:
    return {"cell_type": "markdown", "metadata": {}, "source": _source(text)}


def code(text: str) -> dict:
    return {
        "cell_type": "code",
        "execution_count": None,
        "metadata": {},
        "outputs": [],
        "source": _source(text),
    }


cells = [
    markdown(
        """
        # Raportare articole și citații conform OMEC nr. 3019/2025

        ## Comisia 9 – Inginerie electrică

        Acest notebook conduce întregul flux: Google Scholar, clasificare J/C,
        corelare cu exporturile Web of Science, registre WoS/Google cu intrări
        unice și completarea șablonului de raportare.

        > **Stadiu: testare și validare.** Proiectul poate produce erori,
        > omisiuni, metadate incomplete, asocieri false sau clasificări greșite.
        > Fiecare fișier intermediar este salvat pentru audit. Verificați toate
        > intrările, formulele și punctajele înainte de utilizare oficială.

        Proiectul este conceput în primul rând pentru cercetători cu indice
        Hirsch Google Scholar sub 20. Aceasta este o recomandare operațională,
        nu o regulă SerpApi: h-indexul nu determină direct consumul. Pentru
        profiluri mari sau cu multe citări poate fi necesar un plan plătit.
        La 15 iulie 2026 planul Starter era 25 USD/lună, nu 25 EUR. Verificați
        prețul actual la https://serpapi.com/pricing.
        """
    ),
    markdown(
        """
        ## Fișiere păstrate de proiect

        1. data/intermediate/Articole_Google_neclasificat.xlsx – checkpoint
           Google Scholar, actualizat după fiecare articol;
        2. fișierele log, status.json și summary.json – progres și diagnostic;
        3. data/intermediate/Articole_Google.xlsx – export cu tip J/C;
        4. data/output/Articole_Citari_WoS.xlsx – intrări WoS unice;
        5. data/output/Articole_Citari_Google.xlsx – restul intrărilor Google;
        6. data/output/Criterii_CNATDCU_2025_completat.xlsx – raportul final.

        Fișierele-sursă nu sunt suprascrise. Directoarele de date sunt excluse
        din Git pentru protejarea datelor personale și a materialelor terțe.
        """
    ),
    markdown(
        """
        ## 1. Instalarea Anaconda și Jupyter

        Pe Windows:

        1. descărcați Anaconda Distribution, Windows 64-Bit Graphical
           Installer, de la https://www.anaconda.com/download;
        2. porniți kitul, citiți termenii și alegeți Just Me;
        3. păstrați Create shortcuts și evitați adăugarea permanentă în PATH
           dacă nu este necesară;
        4. finalizați instalarea și deschideți Anaconda Prompt;
        5. descărcați repository-ul cu Git sau Code, Download ZIP și extrageți
           arhiva într-un director propriu;
        6. din rădăcina proiectului rulați:

               conda env create -f environment.yml
               conda activate raportare-omec-3019
               jupyter lab

        Alegeți kernelul Python (raportare OMEC 3019). Dacă lipsește:

               python -m ipykernel install --user --name raportare-omec-3019 --display-name "Python (raportare OMEC 3019)"

        Instrucțiuni oficiale:
        https://www.anaconda.com/docs/getting-started/anaconda/install/windows-gui-install
        și https://jupyter.org/install.
        """
    ),
    code(
        """
        from pathlib import Path
        import sys

        def gaseste_radacina(start: Path) -> Path:
            for candidate in (start.resolve(), *start.resolve().parents):
                if (candidate / "scripts" / "pipeline.py").exists():
                    return candidate
            raise FileNotFoundError(
                "Nu găsesc scripts/pipeline.py. Deschideți notebookul din proiect."
            )

        ROOT = gaseste_radacina(Path.cwd())
        if str(ROOT) not in sys.path:
            sys.path.insert(0, str(ROOT))
        print(f"Rădăcina proiectului: {ROOT}")
        print(f"Python: {sys.version.split()[0]}")
        """
    ),
    markdown(
        """
        ## 2. Instalarea/verificarea bibliotecilor

        Celula următoare verifică dependențele. Dacă unele lipsesc, instalează
        requirements.txt în kernelul curent. După instalare poate fi necesar
        Kernel, Restart Kernel.
        """
    ),
    code(
        """
        import importlib.util
        import subprocess

        module_necesare = [
            "openpyxl", "pandas", "requests", "scholarly", "xlrd", "dotenv"
        ]
        lipsa = [m for m in module_necesare if importlib.util.find_spec(m) is None]
        if lipsa:
            print("Lipsesc:", ", ".join(lipsa))
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "-r", str(ROOT / "requirements.txt")]
            )
            print("Instalare încheiată. Reporniți kernelul dacă este necesar.")
        else:
            print("Toate bibliotecile necesare sunt disponibile.")
        """
    ),
    markdown(
        """
        ## 3. Obținerea cheii SerpApi

        1. creați un cont la https://serpapi.com/users/sign_up;
        2. confirmați e-mailul dacă vi se solicită;
        3. accesați https://serpapi.com/pricing și selectați planul Free;
        4. la 15 iulie 2026 planul Free includea 250 căutări/lună;
        5. din Dashboard sau Your Account copiați cheia API privată;
        6. copiați .env.example ca .env și completați SERPAPI_API_KEY, sau
           introduceți cheia mascat în celula de configurare.

        Nu salvați cheia în notebook, Git, capturi sau issue-uri. Dacă a fost
        publicată, revocați/regenerați cheia.

        Indicele Hirsch nu este unitate de facturare. Sunt folosite cereri
        pentru profil, detaliile articolelor și paginile citărilor. Pentru
        h-index sub 20 cota gratuită poate fi suficientă în multe cazuri, fără
        garanție. Cache-ul reduce repetarea. Planul Starter era 25 USD/lună
        pentru 1.000 de căutări; verificați moneda, taxele și reînnoirea.
        """
    ),
    markdown(
        """
        ## 4. Profilul Google Scholar trebuie să fie public

        1. autentificați-vă pe https://scholar.google.com;
        2. deschideți My profile sau Profilul meu;
        3. completați numele, afilierea, domeniile și e-mailul instituțional;
        4. verificați publicațiile, eliminați asocierile greșite și gestionați
           duplicatele cu atenție;
        5. în editarea profilului activați Make my profile public;
        6. salvați și testați URL-ul într-o fereastră Incognito în care nu
           sunteți autentificat.

        URL-ul trebuie să conțină parametrul user, de exemplu:

            https://scholar.google.com/citations?hl=en&user=IDENTIFICATOR&view_op=list_works

        Ajutor oficial: https://scholar.google.com/intl/us/scholar/help.html.
        """
    ),
    markdown(
        """
        ## 5. Obținerea fișierului Articole_WOS

        1. autentificați-vă pe https://www.e-nformation.ro/ cu un cont
           instituțional autorizat;
        2. accesați Web of Science – Core Collection, InCites Journal Citation
           Reports, Derwent Innovations Index, Clarivate Analytics;
        3. selectați explicit Web of Science Core Collection;
        4. în câmpul Author introduceți numele cercetătorului;
        5. rafinați după variante de nume, afiliere, ORCID, ResearcherID și ani;
        6. verificați manual lista și eliminați autorii omonimi;
        7. selectați Export, Excel;
        8. alegeți Records from 1 to n;
        9. la Record Content alegeți Full Record;
        10. salvați data/input/Articole_WOS.xls sau .xlsx.

        O căutare doar după nume poate include omonimi sau poate omite variante.
        Documentați filtrele și păstrați exportul original.
        """
    ),
    markdown(
        """
        ## 6. Obținerea fișierului Citari_WOS

        1. din setul validat selectați Citation Report în partea de sus;
        2. notați numărul publicațiilor, Times Cited și Citing Articles;
        3. în tabelul de jos, pentru fiecare lucrare faceți clic pe numărul de
           citări din indexul folosit;
        4. în pagina Citing Articles selectați Add to Marked List și All
           records on page;
        5. repetați pe toate paginile și pentru fiecare articol, fără a goli
           Marked List;
        6. la final deschideți Marked List din stânga și fila Documents;
        7. verificați numărul cu Citing Articles distincte din Citation Report;
        8. același document poate cita mai multe lucrări, astfel suma Times
           Cited poate fi mai mare decât lista de documente unice;
        9. selectați Export, Excel, Records from 1 to n, Full Record;
        10. salvați data/input/Citari_WOS.xls sau .xlsx.

        Pentru peste 1.000 de înregistrări pot fi necesare intervale. Reuniți
        exporturile cu un singur antet. Ghid: docs/EXPORT_WEB_OF_SCIENCE.md.
        """
    ),
    markdown(
        """
        ## 7. Configurarea proiectului

        Modificați valorile următoare. Pentru primul test puteți folosi
        MAX_ARTICLES = 2 și MAX_CITATIONS_PER_ARTICLE = 5. Pentru rularea
        completă utilizați None. FORCE_REFRESH = False permite reluarea din
        cache și economisește credite.
        """
    ),
    code(
        """
        import os
        from dotenv import load_dotenv
        from scripts.pipeline import (
            ProjectPaths,
            author_id_from_profile_url,
            check_required_inputs,
            run_classification,
            run_google_scholar,
            run_report,
            run_unique_split,
        )

        load_dotenv(ROOT / ".env")
        PATHS = ProjectPaths(ROOT)
        PATHS.prepare()

        PROFILE_URL = os.getenv(
            "GOOGLE_SCHOLAR_PROFILE_URL",
            "https://scholar.google.com/citations?hl=en&user=INLOCUITI_ID&view_op=list_works",
        )
        CROSSREF_MAILTO = os.getenv("CROSSREF_MAILTO", "")
        BACKEND = "serpapi"
        FORCE_REFRESH = False
        MAX_ARTICLES = None
        MAX_CITATIONS_PER_ARTICLE = None

        ARTICOLE_WOS = PATHS.input_dir / "Articole_WOS.xls"
        CITARI_WOS = PATHS.input_dir / "Citari_WOS.xls"
        SABLON_RAPORTARE = PATHS.input_dir / "Criterii_CNATDCU_2025.xlsx"

        print("Director intrări:", PATHS.input_dir)
        print("Director rezultate:", PATHS.output_dir)
        """
    ),
    code(
        """
        from getpass import getpass

        SERPAPI_API_KEY = os.getenv("SERPAPI_API_KEY", "").strip()
        if not SERPAPI_API_KEY:
            SERPAPI_API_KEY = getpass(
                "Introduceți cheia SerpApi; nu va fi afișată: "
            ).strip()

        if "INLOCUITI_ID" in PROFILE_URL:
            raise ValueError("Înlocuiți PROFILE_URL cu URL-ul profilului public.")
        AUTHOR_ID = author_id_from_profile_url(PROFILE_URL)
        if not SERPAPI_API_KEY:
            raise ValueError("Cheia SerpApi este goală.")

        print(f"Profil configurat: user={AUTHOR_ID}")
        print(
            f"Cheie SerpApi detectată: da, {len(SERPAPI_API_KEY)} caractere; "
            "valoarea nu este afișată"
        )
        for label, path in {
            "Articole WoS": ARTICOLE_WOS,
            "Citări WoS": CITARI_WOS,
            "Șablon": SABLON_RAPORTARE,
        }.items():
            print(f"{label:14s}: {'OK' if path.exists() else 'LIPSEȘTE'} – {path}")
        """
    ),
    markdown(
        """
        ## 8. Extragerea Google Scholar

        Celula creează imediat un Excel valid și îl actualizează prin
        checkpointuri. Fiecare articol este urmat de citările sale. Articolele
        au Arial 10 bold, citările Arial 10 normal.

        Dacă SerpApi nu returnează detalii sau citări pentru un articol,
        proiectul păstrează informațiile disponibile și continuă. Dacă rularea
        este întreruptă, nu ștergeți data/cache; rulați din nou celula.
        """
    ),
    code(
        """
        rezultat_google = run_google_scholar(
            PATHS,
            profile_url=PROFILE_URL,
            serpapi_api_key=SERPAPI_API_KEY,
            crossref_mailto=CROSSREF_MAILTO,
            backend=BACKEND,
            force_refresh=FORCE_REFRESH,
            max_articles=MAX_ARTICLES,
            max_citations_per_article=MAX_CITATIONS_PER_ARTICLE,
        )
        rezultat_google
        """
    ),
    code(
        """
        import json

        status_path = PATHS.google_raw.with_suffix(".status.json")
        log_path = PATHS.google_raw.with_suffix(".log")
        if status_path.exists():
            stare = json.loads(status_path.read_text(encoding="utf-8"))
            print(json.dumps(stare, ensure_ascii=False, indent=2))
        if log_path.exists():
            ultimele = log_path.read_text(encoding="utf-8").splitlines()[-15:]
            print("\\nUltimele mesaje:")
            print("\\n".join(ultimele))
        print("\\nExcel parțial/final:", PATHS.google_raw, PATHS.google_raw.exists())
        """
    ),
    markdown(
        """
        ## 9. Clasificarea J – jurnal / C – conferință

        Clasificarea nu ține cont de litere mari/mici. O denumire devine C dacă
        include conference, symposium, proceedings sau un acronim fără spații
        între paranteze. Orice altă denumire nenulă devine J; lipsa denumirii
        lasă tipul gol.

        Regula este o euristică. Verificați toate rândurile, mai ales revistele
        care conțin Proceedings și denumirile cu paranteze. Corectați manual
        Articole_Google.xlsx, apoi reluați doar etapele următoare.
        """
    ),
    code(
        """
        check_required_inputs(PATHS.google_raw)
        rezultat_clasificare = run_classification(PATHS)
        rezultat_clasificare
        """
    ),
    code(
        """
        from collections import Counter
        from openpyxl import load_workbook

        wb = load_workbook(PATHS.google_classified, read_only=True, data_only=False)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        counts = Counter(str(ws.cell(r, 1).value or "") for r in range(2, ws.max_row + 1))
        titlu_col = headers.index("Titlu") + 1
        venue_col = headers.index("Revista / Conferința") + 1
        lipsa_venue = [
            ws.cell(r, titlu_col).value
            for r in range(2, ws.max_row + 1)
            if not ws.cell(r, venue_col).value
        ]
        wb.close()
        print("Tipuri:", dict(counts))
        print("Rânduri fără revistă/conferință:", len(lipsa_venue))
        print("Exemple de completat:", lipsa_venue[:10])
        """
    ),
    markdown(
        """
        ## 10. Registrele WoS și Google cu intrări unice

        WoS are prioritate. Articolele WoS sunt bold și sunt urmate numai de
        citările lor identificate în Citari_WOS. Rândurile mutate în WoS sunt
        eliminate din Google. Rândurile fără J/C sunt omise și raportate.
        """
    ),
    code(
        """
        check_required_inputs(PATHS.google_classified, ARTICOLE_WOS, CITARI_WOS)
        rezultat_separare = run_unique_split(
            PATHS,
            articole_wos=ARTICOLE_WOS,
            citari_wos=CITARI_WOS,
            fuzzy_threshold=0.90,
        )
        for cheie, valoare in rezultat_separare.items():
            if not isinstance(valoare, (list, dict)):
                print(f"{cheie}: {valoare}")
        print("Verificare:", rezultat_separare.get("verificare"))
        """
    ),
    markdown(
        """
        ## 11. Completarea raportului

        Sunt completate categoriile 2.1.a, 2.1.b, 2.2.a, 2.2.b, 3.1.a și
        3.1.b. În categoriile de citări, articolul-părinte este bold cu fond
        verde și urmat de citări. Un articol fără citări nu este inclus acolo.
        Fiecare rând Google se termină cu Google Scholar.

        Rândurile completate folosesc Arial 10. Formulele și sumele sunt
        translate și validate. Șablonul-sursă nu este suprascris.
        """
    ),
    code(
        """
        check_required_inputs(PATHS.wos_unique, PATHS.google_unique, SABLON_RAPORTARE)
        rezultat_raport = run_report(PATHS, template_path=SABLON_RAPORTARE)
        for cheie, valoare in rezultat_raport.items():
            print(f"{cheie}: {valoare}")
        """
    ),
    code(
        """
        from openpyxl import load_workbook

        rezultate = [PATHS.wos_unique, PATHS.google_unique, PATHS.report]
        for path in rezultate:
            if not path.exists() or path.stat().st_size == 0:
                raise AssertionError(f"Fișier lipsă sau gol: {path}")
            print(f"OK: {path.name} – {path.stat().st_size / 1024:.1f} KB")

        wb = load_workbook(PATHS.report, data_only=False, read_only=True)
        try:
            if not {"PROF", "SINTEZA"}.issubset(wb.sheetnames):
                raise AssertionError(f"Foi neașteptate: {wb.sheetnames}")
            formule = []
            erori = []
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        value = cell.value
                        if isinstance(value, str) and value.startswith("="):
                            formule.append((ws.title, cell.coordinate, value))
                            if "#REF!" in value.upper():
                                erori.append((ws.title, cell.coordinate, value))
            if erori:
                raise AssertionError(f"Referințe #REF!: {erori[:10]}")
        finally:
            wb.close()

        print(f"VALIDARE FINALĂ: OK – {len(formule)} formule fără #REF!.")
        print("Deschideți raportul în Excel, recalculați și verificați manual.")
        """
    ),
    markdown(
        """
        ## 12. Checklist obligatoriu

        - comparați publicațiile cu profilul Google Scholar;
        - verificați J/C pentru fiecare rând;
        - confirmați DOI, ISSN/eISSN, an, volum, număr și pagini;
        - verificați fiecare asociere articol–citare;
        - tratați autocitările conform regulilor aplicabile;
        - confirmați indexarea WoS, IEEE Xplore și BDI;
        - verificați duplicatele după DOI și titlu;
        - deschideți Excelul final și recalculați formulele;
        - comparați punctajele cu ordinul, rectificările și regulile instituției;
        - arhivați intrările, ieșirile, logurile și data rulării.

        ## Declinarea răspunderii

        Proiectul este furnizat ca atare, fără garanții. Autorul/dezvoltatorul
        nu își asumă răspunderea pentru greșeli, omisiuni, costuri API,
        pierderi, respingerea unui dosar sau alte consecințe. Utilizarea nu
        creează nicio obligație legală, contractuală, profesională sau
        fiduciară pentru dezvoltator. Proiectul se folosește pe propriul risc
        și nu constituie consultanță juridică, administrativă sau academică.

        Utilizatorul răspunde de termenii Google, SerpApi, Crossref, Clarivate,
        Web of Science și E-nformation, de protecția datelor și de dreptul de a
        folosi/exporta materialele.

        ## Licență și open source

        Codul este MIT și suplimentar CC BY 4.0. Documentația, notebookul și
        materialele originale sunt CC BY 4.0. Datele și exporturile terților nu
        sunt relicențiate. MIT este licența software principală deoarece
        Creative Commons nu recomandă CC ca licență principală pentru software.

        Consultați README.md, DISCLAIMER.md, PRIVACY.md, NOTICE.md, LICENSE și
        LICENSE-CC-BY-4.0.md.

        Referință oficială OMEC nr. 3019/2025:
        https://legislatie.just.ro/Public/DetaliiDocumentAfis/294663.
        """
    ),
]


notebook = {
    "cells": cells,
    "metadata": {
        "kernelspec": {
            "display_name": "Python (raportare OMEC 3019)",
            "language": "python",
            "name": "python3",
        },
        "language_info": {
            "codemirror_mode": {"name": "ipython", "version": 3},
            "file_extension": ".py",
            "mimetype": "text/x-python",
            "name": "python",
            "nbconvert_exporter": "python",
            "pygments_lexer": "ipython3",
            "version": "3.11",
        },
    },
    "nbformat": 4,
    "nbformat_minor": 5,
}


OUTPUT.write_text(
    json.dumps(notebook, ensure_ascii=False, indent=1) + "\n",
    encoding="utf-8",
)
print(f"Notebook generat: {OUTPUT}")

